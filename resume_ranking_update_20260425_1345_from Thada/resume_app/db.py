"""
db.py — Data layer, theming, and shared utilities.
Olympic Industries PLC — HR Intelligence Platform
"""

import io
import os
import csv
import sys
from pathlib import Path

import psycopg2
import psycopg2.extras
import pandas as pd
import streamlit as st

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# ── Config ─────────────────────────────────────────────────────────────────────

PG_CONN = {
    "host":     os.environ.get("PG_HOST", "localhost"),
    "port":     int(os.environ.get("PG_PORT", "5432")),
    "dbname":   os.environ.get("PG_DBNAME", "resume_ranking"),
    "user":     os.environ.get("PG_USER", "postgres"),
    "password": os.environ.get("PG_PASSWORD", "ai&dt@OIPLC"),
}

# Project root = parent of resume_app/
PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESUMES_BASE = os.environ.get("RESUMES_BASE", str(PROJECT_ROOT / "downloaded_resumes"))
RANKER_PATH  = os.environ.get("RANKER_PATH",  str(PROJECT_ROOT / "ranker.py"))
VENV_PYTHON  = os.environ.get("VENV_PYTHON",  sys.executable)
LOGO_PATH    = str(Path(__file__).resolve().parent / "plc_logo_w_text.png")

DEPARTMENTS = [
    "Finance and Accounts",
    "Admin",
    "AI & Digital Transformation",
    "Brand & Marketing",
    "Corporate Affairs",
    "Customer Service Department (CSD)",
    "Delivery",
    "Distribution",
    "Engineering",
    "Export",
    "External Audit",
    "Factory Administration",
    "Field Force",
    "Human Resource (HR)",
    "Information & Communication Technology (ICT)",
    "Import",
    "Institutional Sales",
    "Internal Audit",
    "Legal Affairs",
    "Local Procurement",
    "Management",
    "Market Audit",
    "Mechanical",
    "Management Information System (MIS)",
    "Operations",
    "Plastic Production",
    "Supply Chain",
    "Production",
    "Quality Assurance Department (QAD)",
    "ERP - SAP",
    "Sales",
    "Secretariat",
    "Security",
    "Share",
    "Store",
    "Transport",
    "VAT / VAT & Delivery",
]

# Authoritative department list used by ranker --department, New Job form,
# assign-dept UI, and nav filters.  "Uncategorized" is the first-class
# fallback for legacy jobs with no department assigned.
DEPARTMENT_LIST = ["Uncategorized"] + DEPARTMENTS

EXPERIENCE_OPTIONS = ["Any","1 year","2 years","3 years","5 years","7 years","10 years","15+ years"]
EDUCATION_OPTIONS  = ["Any","SSC","HSC","Diploma","Bachelor's","Master's","MBA","PhD","Professional Certification"]

COMMON_SKILLS = [
    # Leadership & Management
    "Team Leadership","People Management","Strategic Planning","Change Management",
    "Performance Management","Stakeholder Management","Budget Management",
    "Cross-functional Collaboration","Conflict Resolution","Decision Making",

    # Finance & Accounts
    "Financial Reporting","Budgeting & Forecasting","Cost Control",
    "Accounts Payable","Accounts Receivable","Tax Compliance",
    "Audit","Payroll Management","Treasury Management","VAT Compliance",
    "Tally","QuickBooks",

    # Sales & Marketing
    "Sales Management","Key Account Management","Trade Marketing",
    "Brand Management","Market Research","Digital Marketing",
    "Retail Management","Distribution Management","Field Sales",
    "Institutional Sales","Territory Management","Trade Promotion",

    # Supply Chain & Logistics
    "Supply Chain Management","Logistics","Procurement","Inventory Management",
    "Production Planning","Demand Forecasting","Warehouse Management",
    "Import & Export","Customs Clearance","Fleet Management","Cold Chain",

    # HR
    "Recruitment","Training & Development","HR Policy","Labor Law",
    "Employee Relations","Compensation & Benefits",
    "Organizational Development","Talent Management","HRIS",

    # Production & Quality
    "Production Management","Quality Control","Quality Assurance",
    "Six Sigma","Lean Manufacturing","GMP","HACCP","ISO Standards",
    "Process Improvement","FMCG Manufacturing","Food Safety",

    # Engineering & Maintenance
    "Mechanical Engineering","Electrical Engineering","Preventive Maintenance",
    "Equipment Troubleshooting","AutoCAD","PLC","Industrial Automation",

    # IT & Systems
    "SAP","ERP","Power BI","Excel","SQL","Data Analysis",
    "Process Automation","Python","Machine Learning","Network Administration",
    "IT Support","MIS Reporting","Database Management",

    # Admin & General
    "MS Office","Report Writing","Presentation Skills","Negotiation",
    "Vendor Management","Contract Management","Compliance","Documentation",
    "FMCG Experience","Project Management","Agile",
]


SKILL_DOMAINS = {
    "🤖 AI & Technology": [
        "Machine Learning","Deep Learning","Computer Vision","NLP","Python",
        "Data Analysis","Power BI","Tableau","SQL","Process Automation",
        "RPA","FastAPI","Docker","n8n","ERP Implementation","Digital Strategy",
        "AI Project Management","MIS Reporting","Report Automation","ETL",
        "Database Management","Dashboard Development",
    ],
    "💰 Finance & Audit": [
        "Financial Reporting","Budgeting & Forecasting","Cost Control","Cost Accounting",
        "Accounts Payable","Accounts Receivable","Tax Compliance","VAT Compliance",
        "Audit","Payroll Management","Treasury Management","Cash Flow Management",
        "Bank Reconciliation","Fixed Assets Management","Tally","SAP FICO",
        "Internal Audit","External Audit","Risk Assessment","Financial Controls",
        "Process Audit","Stock Audit","Compliance Audit",
    ],
    "📦 Supply Chain & Logistics": [
        "Supply Chain Management","Inventory Management","Demand Planning",
        "Warehouse Management","S&OP","Logistics Coordination","Stock Reconciliation",
        "Cold Chain Management","3PL Management","Supply Chain Analytics",
        "Fleet Management","Route Planning","Vehicle Maintenance","Driver Management",
        "Delivery Scheduling","Last Mile Delivery","Store Management",
        "FIFO/FEFO Management","Material Handling","GRN Processing","Stock Audit",
    ],
    "🏭 Production & Quality": [
        "Production Planning","Production Management","OEE Improvement",
        "Line Balancing","Shift Management","FMCG Manufacturing",
        "Injection Moulding","Blow Moulding","Plastic Processing",
        "Machine Operation","Downtime Reduction","Batch Production",
        "Quality Control","Quality Assurance","GMP","HACCP","ISO 22000",
        "FSSC 22000","Food Safety","In-process Quality Check",
        "Raw Material Inspection","Finished Goods Testing",
        "Microbiology","Laboratory Management","SOP Development",
    ],
    "⚙️ Engineering & Maintenance": [
        "Mechanical Engineering","Electrical Engineering","Preventive Maintenance",
        "Breakdown Maintenance","Equipment Troubleshooting","AutoCAD",
        "PLC Programming","Pneumatics & Hydraulics","Boiler Operation",
        "Utility Management","Energy Management","Fabrication","Welding",
    ],
    "📊 Sales & Marketing": [
        "Sales Management","Key Account Management","Trade Marketing",
        "Brand Management","Market Research","Digital Marketing",
        "Retail Management","Distribution Management","Field Sales",
        "Institutional Sales","Territory Management","Trade Promotion",
        "Campaign Management","Consumer Insights","Media Planning",
        "ATL/BTL Marketing","Social Media Marketing","Content Creation",
        "Marketing Analytics","New Product Development","Packaging Development",
        "Beat Planning","Outlet Expansion","Market Development",
    ],
    "🚚 Distribution & Field": [
        "Distribution Network Management","Channel Management","Distributor Management",
        "Secondary Sales","Route to Market","Coverage Expansion",
        "Van Sales","Depot Management","Stock Replenishment",
        "Market Survey","Competitor Analysis","Price Monitoring",
        "Outlet Audit","Shelf Share Analysis","Field Data Collection",
        "Market Intelligence","Distribution Audit",
    ],
    "👥 Human Resource": [
        "Recruitment & Selection","Training & Development","HR Policy Development",
        "Performance Management","Compensation & Benefits","Employee Relations",
        "Organizational Development","Talent Management","HRIS","Succession Planning",
        "HR Compliance","Labor Law","Grievance Handling","Onboarding",
    ],
    "🖥️ ICT & MIS": [
        "Network Administration","IT Infrastructure","Cybersecurity",
        "ERP Support","Hardware & Software Troubleshooting","Server Management",
        "Active Directory","IT Helpdesk","Cloud Computing","Backup & Recovery",
        "IT Procurement","CCTV & Access Control","Advanced Excel",
        "KPI Reporting","Python",
    ],
    "🔵 SAP & ERP": [
        "SAP MM","SAP SD","SAP FI","SAP CO","SAP PP","SAP QM","SAP WM",
        "SAP HR","SAP BASIS","SAP ABAP","S/4HANA","SAP Implementation",
        "SAP Configuration","SAP Reporting","SAP Integration",
    ],
    "🌐 Export & Import": [
        "Export Documentation","LC Management","Customs Clearance",
        "BGMEA/BEPZA Compliance","Freight Forwarding","Incoterms",
        "Shipping & Logistics","Import Documentation","Bond Licensing",
        "Foreign Currency Management","Trade Finance",
    ],
    "📋 Admin & Compliance": [
        "Office Management","Facility Management","Vendor Management",
        "Contract Management","Record Keeping","Travel & Logistics Coordination",
        "Event Management","Document Control","Scheduling & Calendar Management",
        "Corporate Governance","Regulatory Compliance","Legal Documentation",
        "Contract Drafting & Review","Company Secretarial Work","RJSC Compliance",
        "BIDA Compliance","Intellectual Property","Due Diligence",
        "Factory Compliance","Industrial Relations","Factory Act Compliance",
    ],
    "🛡️ Security & Store": [
        "Physical Security","Access Control","CCTV Monitoring",
        "Security Risk Assessment","Crisis Management","Guard Management",
        "Loss Prevention","Inventory Control","Bin Card Management",
        "Warehouse Layout Optimization",
    ],
    "🏢 Leadership & Strategy": [
        "Team Leadership","People Management","Strategic Planning","Change Management",
        "Performance Management","Stakeholder Management","Budget Management",
        "Cross-functional Collaboration","Conflict Resolution","Decision Making",
        "Operations Management","Project Management","Business Process Improvement",
        "KPI Management","Agile","Six Sigma",
    ],
    "📝 General & Communication": [
        "Report Writing","Presentation Skills","Negotiation","Problem Solving",
        "Data Entry","Filing & Documentation","FMCG Experience",
        "Communication Skills","Team Coordination","MS Office",
    ],
}

# ── Department Skill Profiles (Part 2) ─────────────────────────────────────────
# Used by ranker.py to inject department-specific scoring context into the prompt.
# Each entry defines:
#   "core_skills"     : must-have tools/skills. Missing >50% caps skills_score at 50.
#   "bonus_skills"    : differentiators that push score above 75.
#   "anti_skills"     : skills that are irrelevant and should not inflate score.
#   "scoring_note"    : free-text guidance injected directly into LLM prompt.

DEPARTMENT_SKILL_PROFILES = {

    "Brand & Marketing": {
        "core_skills": [
            "Adobe Premiere Pro", "Adobe After Effects", "Adobe Photoshop",
            "Canva", "Video Editing", "Graphic Design", "Content Creation",
            "Social Media Marketing", "Brand Management", "Campaign Management",
            "Digital Marketing", "Market Research", "Consumer Insights",
            "Copywriting", "ATL/BTL Marketing",
        ],
        "bonus_skills": [
            "Adobe Illustrator", "Motion Graphics", "Media Planning",
            "Google Analytics", "Meta Ads Manager", "Marketing Analytics",
            "New Product Development", "Packaging Design", "Trade Marketing",
        ],
        "anti_skills": ["PLC Programming", "Boiler Operation", "Tax Compliance"],
        "scoring_note": (
            "For Brand & Marketing / Design roles: heavy weight on creative tools "
            "(Adobe suite, video editing). Candidates without Adobe Premiere OR "
            "After Effects OR equivalent video/design tools score a maximum of 55 "
            "on skills_score for any Designer or Content role. "
            "Social media growth metrics and portfolio evidence are strong signals."
        ),
    },

    "AI & Digital Transformation": {
        "core_skills": [
            "Python", "Machine Learning", "Data Analysis", "SQL",
            "Power BI", "ERP Implementation", "Process Automation",
            "Digital Strategy", "AI Project Management", "Dashboard Development",
        ],
        "bonus_skills": [
            "Deep Learning", "NLP", "Computer Vision", "Docker",
            "FastAPI", "n8n", "RPA", "ETL", "Tableau", "MIS Reporting",
            "S/4HANA", "SAP Integration", "Change Management",
        ],
        "anti_skills": ["Manual Data Entry", "Physical Security"],
        "scoring_note": (
            "For AI & Digital Transformation roles: Python proficiency and at least "
            "one deployed ML/AI project are strong positive signals. "
            "ERP (SAP) implementation experience is highly valued in FMCG context. "
            "Candidates with only theoretical knowledge and no production deployments "
            "should score no higher than 65 on skills_score."
        ),
    },

    "Finance and Accounts": {
        "core_skills": [
            "Financial Reporting", "Budgeting & Forecasting", "Cost Control",
            "Accounts Payable", "Accounts Receivable", "Tax Compliance",
            "VAT Compliance", "Tally", "SAP FICO", "Bank Reconciliation",
            "Fixed Assets Management", "Payroll Management",
        ],
        "bonus_skills": [
            "Treasury Management", "Cash Flow Management", "Cost Accounting",
            "Internal Audit", "External Audit", "Risk Assessment",
            "Financial Controls", "IFRS", "Bangladesh Tax Law", "NBFI Regulations",
        ],
        "anti_skills": ["Video Editing", "CAD", "Machine Operation"],
        "scoring_note": (
            "For Finance roles: CA / ACCA / CPA qualification is a strong differentiator "
            "and should push education_score upward. VAT Compliance and Bangladesh "
            "Tax Law knowledge are essential for senior roles. SAP FICO is a bonus "
            "given Olympic's SAP environment."
        ),
    },

    "Human Resource (HR)": {
        "core_skills": [
            "Recruitment & Selection", "Training & Development", "HR Policy Development",
            "Performance Management", "Compensation & Benefits", "Labor Law",
            "Employee Relations", "HRIS", "Onboarding", "HR Compliance",
        ],
        "bonus_skills": [
            "Organizational Development", "Talent Management", "Succession Planning",
            "Grievance Handling", "Bangladesh Labor Law", "SAP HR",
            "Learning Management System (LMS)", "Employer Branding",
        ],
        "anti_skills": ["Mechanical Engineering", "PLC Programming", "Tax Compliance"],
        "scoring_note": (
            "For HR roles: knowledge of Bangladesh Labour Act 2006 is mandatory "
            "for mid/senior roles. Candidates with PGDHRM or CIPD qualification "
            "should receive a bonus on education_score. Factory HR experience "
            "is a differentiator in the FMCG manufacturing context."
        ),
    },

    "Information & Communication Technology (ICT)": {
        "core_skills": [
            "Network Administration", "IT Infrastructure", "Server Management",
            "Hardware & Software Troubleshooting", "Active Directory",
            "IT Helpdesk", "ERP Support", "Backup & Recovery",
        ],
        "bonus_skills": [
            "Cybersecurity", "Cloud Computing", "CCTV & Access Control",
            "Firewall Management", "VMware / Hyper-V", "IT Procurement",
            "SAP BASIS", "Python", "Advanced Excel",
        ],
        "anti_skills": ["Video Editing", "Brand Management", "Quality Control"],
        "scoring_note": (
            "For ICT roles: CCNA, CompTIA, or Microsoft certifications are strong "
            "education differentiators. Candidates supporting a large enterprise "
            "(500+ users) score higher on experience_score."
        ),
    },

    "ERP - SAP": {
        "core_skills": [
            "SAP MM", "SAP SD", "SAP FI", "SAP CO", "SAP PP",
            "SAP Configuration", "SAP Reporting", "S/4HANA",
            "SAP Implementation", "SAP Integration",
        ],
        "bonus_skills": [
            "SAP QM", "SAP WM", "SAP HR", "SAP ABAP", "SAP BASIS",
            "SAP Fiori", "SAP BW/BI", "End-User Training",
        ],
        "anti_skills": ["Manual Bookkeeping", "Physical Security", "Machine Operation"],
        "scoring_note": (
            "For SAP roles: at least 2 full-cycle SAP implementations are expected "
            "at the senior level. Module depth matters more than breadth. "
            "S/4HANA experience is a strong differentiator for Olympic's current roadmap."
        ),
    },

    "Supply Chain": {
        "core_skills": [
            "Supply Chain Management", "Inventory Management", "Demand Planning",
            "Warehouse Management", "S&OP", "Procurement", "Vendor Management",
            "Import & Export", "Logistics Coordination", "Stock Reconciliation",
        ],
        "bonus_skills": [
            "Cold Chain Management", "3PL Management", "Supply Chain Analytics",
            "FIFO/FEFO Management", "SAP MM", "Customs Clearance",
            "Demand Forecasting", "Cost Reduction Initiatives",
        ],
        "anti_skills": ["Video Editing", "HR Policy", "Tax Compliance"],
        "scoring_note": (
            "FMCG supply chain experience is a strong differentiator. "
            "Candidates who have managed high-SKU environments (200+ SKUs) "
            "or nationwide distribution networks score higher on experience_score."
        ),
    },

    "Sales": {
        "core_skills": [
            "Sales Management", "Key Account Management", "Trade Marketing",
            "Retail Management", "Distribution Management", "Field Sales",
            "Territory Management", "Beat Planning", "Outlet Expansion",
        ],
        "bonus_skills": [
            "Institutional Sales", "Trade Promotion", "Market Development",
            "Secondary Sales", "Van Sales", "Distributor Management",
            "CRM", "Sales Analytics", "Channel Management",
        ],
        "anti_skills": ["Software Development", "Mechanical Engineering", "Audit"],
        "scoring_note": (
            "For Sales roles: quantified achievements (revenue growth %, territory "
            "expansion, new outlet counts) are strong positive evidence and should "
            "increase experience_score. FMCG field sales experience is weighted heavily."
        ),
    },

    "Production": {
        "core_skills": [
            "Production Management", "Production Planning", "FMCG Manufacturing",
            "Shift Management", "OEE Improvement", "Line Balancing",
            "Batch Production", "Downtime Reduction", "Machine Operation",
        ],
        "bonus_skills": [
            "Lean Manufacturing", "Six Sigma", "Injection Moulding",
            "Blow Moulding", "Plastic Processing", "SOP Development",
            "5S Methodology", "TPM",
        ],
        "anti_skills": ["Digital Marketing", "Tax Compliance", "Network Administration"],
        "scoring_note": (
            "For Production roles: experience in high-volume FMCG or plastic "
            "manufacturing lines is essential. Lean / Six Sigma certification "
            "should push education_score upward."
        ),
    },

    "Quality Assurance Department (QAD)": {
        "core_skills": [
            "Quality Control", "Quality Assurance", "GMP", "HACCP",
            "ISO 22000", "Food Safety", "In-process Quality Check",
            "Raw Material Inspection", "Finished Goods Testing",
            "Laboratory Management", "SOP Development",
        ],
        "bonus_skills": [
            "FSSC 22000", "Microbiology", "Halal Certification",
            "ISO 9001", "ISO 14001", "Six Sigma", "Root Cause Analysis",
            "Sensory Evaluation", "Statistical Process Control (SPC)",
        ],
        "anti_skills": ["Software Development", "HR Policy", "Finance Reporting"],
        "scoring_note": (
            "For QAD roles: BSc/MSc in Food Science, Chemistry, Microbiology, "
            "or Chemical Engineering is ideal. HACCP Lead Auditor certification "
            "is a strong differentiator. Practical lab experience in FMCG is required."
        ),
    },

    "Engineering": {
        "core_skills": [
            "Mechanical Engineering", "Electrical Engineering",
            "Preventive Maintenance", "Breakdown Maintenance",
            "Equipment Troubleshooting", "PLC Programming",
            "Pneumatics & Hydraulics", "Utility Management",
        ],
        "bonus_skills": [
            "AutoCAD", "Boiler Operation", "Energy Management",
            "Industrial Automation", "Fabrication", "Welding",
            "HVAC", "Instrumentation",
        ],
        "anti_skills": ["Digital Marketing", "HR Recruitment", "Tax Filing"],
        "scoring_note": (
            "For Engineering/Maintenance roles: practical hands-on experience "
            "with industrial machinery is mandatory. BUET or equivalent engineering "
            "degree is highly valued. Candidates with no factory floor experience "
            "should score no higher than 55 on experience_score."
        ),
    },

    "Distribution": {
        "core_skills": [
            "Distribution Network Management", "Channel Management",
            "Distributor Management", "Secondary Sales", "Route to Market",
            "Coverage Expansion", "Depot Management", "Stock Replenishment",
        ],
        "bonus_skills": [
            "Van Sales", "Fleet Management", "Route Planning",
            "Market Survey", "Outlet Audit", "Distribution Audit",
            "Market Intelligence", "Field Data Collection",
        ],
        "anti_skills": ["Software Engineering", "Audit", "Legal"],
        "scoring_note": (
            "Distribution roles require deep field knowledge of Bangladesh's "
            "district and upazilla-level trade channels. National coverage "
            "management experience is a strong differentiator."
        ),
    },

    "Internal Audit": {
        "core_skills": [
            "Internal Audit", "Process Audit", "Stock Audit",
            "Compliance Audit", "Risk Assessment", "Financial Controls",
            "Audit Planning", "Audit Reporting",
        ],
        "bonus_skills": [
            "CIA (Certified Internal Auditor)", "ACCA", "SAP Audit Trails",
            "Data Analytics for Audit", "Forensic Accounting",
            "ERP Audit", "Bangladesh Tax Law",
        ],
        "anti_skills": ["Machine Operation", "Field Sales", "Network Administration"],
        "scoring_note": (
            "For Internal Audit: CIA or ACCA qualifications are strong education "
            "differentiators. Experience auditing FMCG or manufacturing operations "
            "is preferred over pure financial audit."
        ),
    },

    "Legal Affairs": {
        "core_skills": [
            "Contract Drafting & Review", "Legal Documentation",
            "Corporate Governance", "RJSC Compliance", "Regulatory Compliance",
            "Intellectual Property", "Due Diligence", "Labor Law",
        ],
        "bonus_skills": [
            "BIDA Compliance", "Company Secretarial Work",
            "Industrial Relations", "Factory Act Compliance",
            "Competition Law", "Litigation Management",
        ],
        "anti_skills": ["Machine Operation", "Video Editing", "Supply Chain"],
        "scoring_note": (
            "LLB from a reputable university plus Bar Council enrollment is essential. "
            "Experience advising FMCG or manufacturing companies is preferred. "
            "Knowledge of Bangladesh Companies Act 1994 and Labour Act 2006 is key."
        ),
    },

    "Admin": {
        "core_skills": [
            "Office Management", "Facility Management", "Vendor Management",
            "Document Control", "Scheduling & Calendar Management",
            "Travel & Logistics Coordination", "MS Office",
        ],
        "bonus_skills": [
            "Contract Management", "Event Management", "Record Keeping",
            "ERP Support", "Advanced Excel", "Reporting",
        ],
        "anti_skills": ["PLC Programming", "Machine Learning", "Audit"],
        "scoring_note": (
            "Admin roles value organisational reliability and multi-tasking. "
            "Candidates who have managed large offices (100+ staff) or "
            "multi-location operations score higher on experience_score."
        ),
    },

    "Management Information System (MIS)": {
        "core_skills": [
            "MIS Reporting", "Advanced Excel", "KPI Reporting",
            "Power BI", "Data Analysis", "Dashboard Development",
            "Report Automation", "Database Management",
        ],
        "bonus_skills": [
            "SQL", "Python", "Tableau", "SAP Reporting",
            "ETL Processes", "VBA / Macros",
        ],
        "anti_skills": ["Video Editing", "Machine Operation", "HR Policy"],
        "scoring_note": (
            "MIS roles require strong Excel and BI tool proficiency. "
            "Experience building real-time KPI dashboards for senior management "
            "is a strong differentiator."
        ),
    },
}

# Departments without a specific profile fall back to generic scoring.

RED_FLAG_PRESETS = [
    # Stability
    "Frequent job changes",
    "Short tenures (<1yr)",
    "Employment gaps",
    "Career regression",
    "Over-reliance on one employer",

    # Qualification
    "No relevant degree",
    "Missing certifications",
    "Overqualified",
    "Underqualified",
    "Skills mismatch",

    # Experience
    "No FMCG experience",
    "No industry experience",
    "No leadership experience",
    "No measurable achievements",
    "Vague job descriptions",
    "Limited local market knowledge",

    # Red signals
    "No SAP/ERP exposure",
    "No references listed",
    "Salary expectations too high",
    "Relocation required",
    "Poor communication indicators",
]

SCORE_DIMS = [
    ("overall_score",     "Overall"),
    ("skills_score",      "Skills"),
    ("experience_score",  "Experience"),
    ("leadership_score",  "Leadership"),
    ("education_score",   "Education"),
    ("culture_fit_score", "Culture Fit"),
]

SCORE_COLS = SCORE_DIMS  # legacy alias

VERDICT_CFG = {
    "Shortlist": {"color":"#16A34A","bg":"#DCFCE7","icon":"🟢","dark_bg":"#14532D","dark_color":"#86EFAC"},
    "Maybe":     {"color":"#D97706","bg":"#FEF3C7","icon":"🟡","dark_bg":"#451A03","dark_color":"#FCD34D"},
    "Reject":    {"color":"#DC2626","bg":"#FEE2E2","icon":"🔴","dark_bg":"#450A0A","dark_color":"#FCA5A5"},
}

VERDICT_STYLE = {
    k: {"icon": v["icon"], "color_dark": v["dark_color"], "color_light": v["color"]}
    for k, v in VERDICT_CFG.items()
}

RED      = "#C8102E"
RED_DARK = "#A00D24"
RED_PALE = "#FEE2E2"

# ── Schema ─────────────────────────────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS jobs (
    id                SERIAL PRIMARY KEY,
    job_label         TEXT UNIQUE NOT NULL,
    job_title         TEXT,
    department        TEXT,
    jd_text           TEXT,
    required_skills   TEXT[],
    red_flags         TEXT[],
    min_experience    TEXT,
    education_req     TEXT,
    weight_skills     INTEGER DEFAULT 50,
    weight_exp        INTEGER DEFAULT 30,
    weight_edu        INTEGER DEFAULT 20,
    weight_leadership INTEGER DEFAULT 10,
    weight_culture    INTEGER DEFAULT 5,
    interviewer_notes TEXT,
    status            TEXT DEFAULT 'Pending',
    created_at        TIMESTAMP DEFAULT NOW(),
    last_ranked_at    TIMESTAMP
);

CREATE TABLE IF NOT EXISTS candidates (
    id                SERIAL PRIMARY KEY,
    job_label         TEXT NOT NULL,
    apply_id          TEXT NOT NULL,
    candidate_name    TEXT,
    email             TEXT,
    mobile            TEXT,
    location          TEXT,
    degree            TEXT,
    university        TEXT,
    experience_detail TEXT,
    age               NUMERIC(5,1),
    expected_salary   TEXT,
    current_salary    TEXT,
    application_date  TEXT,
    bdjobs_score      TEXT,
    has_uploaded_cv   BOOLEAN DEFAULT FALSE,
    profile_txt_path  TEXT,
    pdf_path          TEXT,
    pdf_text_chars    INTEGER DEFAULT 0,
    jd_used           TEXT DEFAULT '',
    overall_score     INTEGER,
    skills_score      INTEGER,
    experience_score  INTEGER,
    leadership_score  INTEGER,
    education_score   INTEGER,
    edu_tier_score    INTEGER,
    edu_degree_score  INTEGER,
    edu_gpa_score     INTEGER,
    culture_fit_score INTEGER,
    experience_years  NUMERIC(4,1),
    strengths         TEXT[],
    gaps              TEXT[],
    risk_flags        TEXT[],
    recommendation    TEXT CHECK (recommendation IN ('Shortlist','Maybe','Reject')),
    reasoning         TEXT,
    hr_override       TEXT,
    hr_note           TEXT,
    ranked_at         TIMESTAMP DEFAULT NOW(),
    rank_error        TEXT,
    UNIQUE (job_label, apply_id)
);

-- HR override audit log (Part 7.5) — every time HR changes a recommendation,
-- a row is appended here.  Required for compliance / fairness audit trails.
CREATE TABLE IF NOT EXISTS hr_audit_log (
    id          SERIAL PRIMARY KEY,
    job_label   TEXT NOT NULL,
    apply_id    TEXT NOT NULL,
    hr_user     TEXT,
    old_value   TEXT,
    new_value   TEXT,
    note        TEXT,
    changed_at  TIMESTAMP DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_c_job   ON candidates(job_label);
CREATE INDEX IF NOT EXISTS idx_c_rec   ON candidates(recommendation);
CREATE INDEX IF NOT EXISTS idx_c_score ON candidates(overall_score DESC);
CREATE INDEX IF NOT EXISTS idx_c_email ON candidates(email);
CREATE INDEX IF NOT EXISTS idx_c_mobile ON candidates(mobile);
CREATE INDEX IF NOT EXISTS idx_j_label ON jobs(job_label);
CREATE INDEX IF NOT EXISTS idx_j_dept  ON jobs(department);
CREATE INDEX IF NOT EXISTS idx_hr_audit_job ON hr_audit_log(job_label);
CREATE INDEX IF NOT EXISTS idx_hr_audit_app ON hr_audit_log(apply_id);
"""

# Additive migrations that run every startup — safe to re-run.
MIGRATION_SQL = """
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS updated_at        TIMESTAMP DEFAULT NOW();
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS weight_leadership INTEGER DEFAULT 10;
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS weight_culture    INTEGER DEFAULT 5;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS edu_tier_score   INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS edu_degree_score INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS edu_gpa_score    INTEGER;
UPDATE jobs SET department = 'Uncategorized'
  WHERE department IS NULL OR department = '';
"""

# ── Connection ─────────────────────────────────────────────────────────────────

def _new_conn():
    """Create a fresh connection and ensure schema + migrations + backfill."""
    conn = psycopg2.connect(**PG_CONN)
    conn.autocommit = True
    ensure_schema(conn)
    return conn


def ensure_schema(conn) -> None:
    """Create tables, run migrations, and backfill `jobs` rows from any
    candidate `job_label` not yet registered.  Idempotent and cheap.
    """
    with conn.cursor() as cur:
        cur.execute(SCHEMA_SQL)
        cur.execute(MIGRATION_SQL)
        # Backfill: any candidate.job_label with no jobs row → Uncategorized.
        cur.execute("""
            INSERT INTO jobs (job_label, department)
            SELECT DISTINCT c.job_label, 'Uncategorized'
              FROM candidates c
              LEFT JOIN jobs j ON j.job_label = c.job_label
             WHERE j.job_label IS NULL
            ON CONFLICT (job_label) DO NOTHING
        """)


def migrate_existing_jobs(conn) -> int:
    """Explicit wrapper for the backfill — returns rows inserted."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO jobs (job_label, department)
            SELECT DISTINCT c.job_label, 'Uncategorized'
              FROM candidates c
              LEFT JOIN jobs j ON j.job_label = c.job_label
             WHERE j.job_label IS NULL
            ON CONFLICT (job_label) DO NOTHING
            RETURNING job_label
        """)
        return len(cur.fetchall())


def get_conn():
    """Return a live psycopg2 connection, reconnecting if stale/closed.

    psycopg2 connections held in @st.cache_resource go stale when Postgres
    restarts or when the session times out. We keep one per-session and
    transparently reconnect on failure.
    """
    try:
        conn = st.session_state.get("pg_conn")
        if conn is None or getattr(conn, "closed", 1):
            conn = _new_conn()
            st.session_state["pg_conn"] = conn
        else:
            # Liveness ping — dead connections raise here.
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
        return conn
    except Exception:
        conn = _new_conn()
        st.session_state["pg_conn"] = conn
        return conn


def fresh_conn():
    """Always returns a brand-new autocommit connection. Caller closes it."""
    return _new_conn()

# ── Metadata ingestion ─────────────────────────────────────────────────────────

def ingest_metadata(job_label: str, meta_csv: str) -> tuple:
    if not os.path.exists(meta_csv):
        return 0, 0
    conn = fresh_conn()
    updated = skipped = 0
    with open(meta_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            apply_id = str(row.get("apply_id") or row.get("ApplyID") or "").strip()
            if not apply_id:
                skipped += 1
                continue
            pdf_file = str(row.get("uploaded_cv_file") or "").strip()
            pdf_path = os.path.join(RESUMES_BASE, job_label, "uploaded_cvs", pdf_file) if pdf_file else ""
            try:
                age = float(str(row.get("age") or "0").replace(",",""))
            except ValueError:
                age = None
            has_cv = str(row.get("has_uploaded_cv") or "").strip().lower() in ("yes","true","1")
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO candidates
                        (job_label, apply_id, candidate_name, email, mobile, location,
                         degree, university, experience_detail, age,
                         expected_salary, current_salary, application_date,
                         bdjobs_score, has_uploaded_cv, pdf_path)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (job_label, apply_id) DO UPDATE SET
                        candidate_name    = EXCLUDED.candidate_name,
                        email             = EXCLUDED.email,
                        mobile            = EXCLUDED.mobile,
                        location          = EXCLUDED.location,
                        degree            = EXCLUDED.degree,
                        university        = EXCLUDED.university,
                        experience_detail = EXCLUDED.experience_detail,
                        age               = EXCLUDED.age,
                        expected_salary   = EXCLUDED.expected_salary,
                        current_salary    = EXCLUDED.current_salary,
                        application_date  = EXCLUDED.application_date,
                        bdjobs_score      = EXCLUDED.bdjobs_score,
                        has_uploaded_cv   = EXCLUDED.has_uploaded_cv,
                        pdf_path          = COALESCE(EXCLUDED.pdf_path, candidates.pdf_path)
                """, (
                    job_label, apply_id,
                    str(row.get("candidate_name") or "").strip(),
                    str(row.get("email") or "").strip(),
                    str(row.get("mobile") or "").strip(),
                    str(row.get("location") or "").strip(),
                    str(row.get("degree") or "").strip(),
                    str(row.get("university") or "").strip(),
                    str(row.get("experience") or "").strip(),
                    age,
                    str(row.get("expected_salary") or "").strip(),
                    str(row.get("current_salary") or "").strip(),
                    str(row.get("application_date") or "").strip(),
                    str(row.get("bdjobs_match_score") or "").strip(),
                    has_cv, pdf_path,
                ))
            updated += 1
    conn.close()
    return updated, skipped

# ── Job queries ────────────────────────────────────────────────────────────────

def fetch_all_jobs(conn) -> pd.DataFrame:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                j.job_label, j.job_title, j.department, j.status,
                j.created_at, j.last_ranked_at,
                COUNT(c.id)                                                   AS total,
                SUM(CASE WHEN c.overall_score IS NOT NULL THEN 1 ELSE 0 END) AS ranked,
                SUM(CASE WHEN c.recommendation='Shortlist' THEN 1 ELSE 0 END) AS shortlisted,
                ROUND(AVG(c.overall_score))                                   AS avg_score
            FROM jobs j
            LEFT JOIN candidates c ON c.job_label = j.job_label
            GROUP BY j.job_label, j.job_title, j.department, j.status,
                     j.created_at, j.last_ranked_at
            ORDER BY j.created_at DESC
        """)
        rows = cur.fetchall()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["created_at"]     = pd.to_datetime(df["created_at"]).dt.strftime("%d %b %Y")
    df["last_ranked_at"] = pd.to_datetime(df["last_ranked_at"]).dt.strftime("%d %b %Y %H:%M")
    return df


def fetch_job_labels(conn) -> list:
    """Back-compat helper used by older pages.  Now also returns department.

    Tuple layout: (job_label, total, ranked, last_ranked_at, department).
    """
    df = fetch_all_jobs(conn)
    if df.empty:
        return []
    result = []
    for _, row in df.iterrows():
        result.append((
            row["job_label"],
            int(row.get("total") or 0),
            int(row.get("ranked") or 0),
            row.get("last_ranked_at"),
            str(row.get("department") or "Uncategorized"),
        ))
    return result


# ── Department queries ─────────────────────────────────────────────────────────

def fetch_departments(conn) -> list:
    """Departments that have at least one candidate (ranked or not).

    Each dict: {department, job_count, total_candidates, ranked_candidates,
    shortlist, maybe, reject, last_run}.  Sorted by ranked_candidates desc.
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                COALESCE(j.department, 'Uncategorized')               AS department,
                COUNT(DISTINCT j.job_label)                           AS job_count,
                COUNT(c.id)                                           AS total_candidates,
                SUM(CASE WHEN c.overall_score IS NOT NULL
                         THEN 1 ELSE 0 END)                           AS ranked_candidates,
                SUM(CASE WHEN c.recommendation='Shortlist'
                         THEN 1 ELSE 0 END)                           AS shortlist,
                SUM(CASE WHEN c.recommendation='Maybe'
                         THEN 1 ELSE 0 END)                           AS maybe,
                SUM(CASE WHEN c.recommendation='Reject'
                         THEN 1 ELSE 0 END)                           AS reject,
                MAX(c.ranked_at)                                      AS last_run
            FROM jobs j
            LEFT JOIN candidates c ON c.job_label = j.job_label
            GROUP BY COALESCE(j.department, 'Uncategorized')
            HAVING COUNT(c.id) > 0
            ORDER BY ranked_candidates DESC NULLS LAST, total_candidates DESC
        """)
        rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get("last_run"):
            d["last_run"] = pd.to_datetime(d["last_run"]).strftime("%d %b %Y %H:%M")
        for k in ("job_count", "total_candidates", "ranked_candidates",
                  "shortlist", "maybe", "reject"):
            d[k] = int(d.get(k) or 0)
        result.append(d)
    return result


def fetch_jobs_by_department(conn, department: str) -> list:
    """All jobs under a department with per-verdict counts."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                j.job_label, j.job_title,
                COALESCE(j.department, 'Uncategorized')               AS department,
                COUNT(c.id)                                           AS total,
                SUM(CASE WHEN c.overall_score IS NOT NULL
                         THEN 1 ELSE 0 END)                           AS ranked,
                SUM(CASE WHEN c.recommendation='Shortlist'
                         THEN 1 ELSE 0 END)                           AS shortlist,
                SUM(CASE WHEN c.recommendation='Maybe'
                         THEN 1 ELSE 0 END)                           AS maybe,
                SUM(CASE WHEN c.recommendation='Reject'
                         THEN 1 ELSE 0 END)                           AS reject,
                MAX(c.ranked_at)                                      AS last_run
            FROM jobs j
            LEFT JOIN candidates c ON c.job_label = j.job_label
            WHERE COALESCE(j.department, 'Uncategorized') = %s
            GROUP BY j.job_label, j.job_title, j.department
            ORDER BY last_run DESC NULLS LAST
        """, (department,))
        rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get("last_run"):
            d["last_run"] = pd.to_datetime(d["last_run"]).strftime("%d %b %Y %H:%M")
        for k in ("total", "ranked", "shortlist", "maybe", "reject"):
            d[k] = int(d.get(k) or 0)
        result.append(d)
    return result


def fetch_candidates_by_department(conn, department: str) -> pd.DataFrame:
    """All ranked candidates in a department, ordered by overall_score desc.

    Adds a sequential 1..N `rank` column after ordering.
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                c.apply_id, c.candidate_name, c.job_label,
                COALESCE(j.department, 'Uncategorized')     AS department,
                j.job_title,
                c.email, c.mobile, c.location,
                c.degree, c.university, c.experience_detail,
                c.age, c.expected_salary, c.current_salary,
                c.application_date,
                c.bdjobs_score, c.has_uploaded_cv, c.pdf_path,
                c.overall_score, c.skills_score, c.experience_score,
                c.leadership_score, c.education_score, c.culture_fit_score,
                c.experience_years, c.recommendation,
                c.reasoning, c.strengths, c.gaps, c.risk_flags,
                c.hr_override, c.hr_note,
                c.pdf_text_chars, c.jd_used, c.rank_error, c.ranked_at
            FROM candidates c
            JOIN jobs j ON j.job_label = c.job_label
            WHERE COALESCE(j.department, 'Uncategorized') = %s
              AND c.overall_score IS NOT NULL
            ORDER BY c.overall_score DESC NULLS LAST,
                     c.ranked_at DESC NULLS LAST
        """, (department,))
        rows = cur.fetchall()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df.insert(0, "rank", range(1, len(df) + 1))
    df["experience_years"] = pd.to_numeric(df["experience_years"], errors="coerce")
    df["age"]              = pd.to_numeric(df["age"], errors="coerce")
    df["ranked_at"]        = pd.to_datetime(df["ranked_at"]).dt.strftime("%d %b %Y %H:%M")
    return df


def get_job_department(conn, job_label: str) -> str:
    """Small helper — returns department for a job_label (or 'Uncategorized')."""
    with conn.cursor() as cur:
        cur.execute("SELECT department FROM jobs WHERE job_label = %s", (job_label,))
        row = cur.fetchone()
    return (row[0] if row and row[0] else "Uncategorized")


def set_job_department(conn, job_label: str, department: str) -> None:
    """Upsert department for a job_label. Used by the assign-dept UI and ranker."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO jobs (job_label, department)
            VALUES (%s, %s)
            ON CONFLICT (job_label) DO UPDATE SET
                department = EXCLUDED.department,
                updated_at = NOW()
        """, (job_label, department))


def fetch_job(conn, job_label: str) -> dict:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT * FROM jobs WHERE job_label = %s", (job_label,))
        row = cur.fetchone()
    return dict(row) if row else {}


def create_job(job_data: dict):
    conn = fresh_conn()
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO jobs
                (job_label, job_title, department, jd_text, required_skills,
                 red_flags, min_experience, education_req,
                 weight_skills, weight_exp, weight_edu,
                 weight_leadership, weight_culture,
                 interviewer_notes, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pending')
            ON CONFLICT (job_label) DO UPDATE SET
                job_title         = EXCLUDED.job_title,
                department        = EXCLUDED.department,
                jd_text           = EXCLUDED.jd_text,
                required_skills   = EXCLUDED.required_skills,
                red_flags         = EXCLUDED.red_flags,
                min_experience    = EXCLUDED.min_experience,
                education_req     = EXCLUDED.education_req,
                weight_skills     = EXCLUDED.weight_skills,
                weight_exp        = EXCLUDED.weight_exp,
                weight_edu        = EXCLUDED.weight_edu,
                weight_leadership = EXCLUDED.weight_leadership,
                weight_culture    = EXCLUDED.weight_culture,
                interviewer_notes = EXCLUDED.interviewer_notes,
                status            = 'Pending'
        """, (
            job_data["job_label"],       job_data["job_title"],
            job_data["department"],      job_data["jd_text"],
            job_data["required_skills"], job_data["red_flags"],
            job_data["min_experience"],  job_data["education_req"],
            job_data["weight_skills"],   job_data["weight_exp"],
            job_data["weight_edu"],
            job_data.get("weight_leadership", 10),
            job_data.get("weight_culture",    5),
            job_data["interviewer_notes"],
        ))
    conn.close()


def update_job_status(job_label: str, status: str):
    conn = fresh_conn()
    with conn.cursor() as cur:
        if status == "Complete":
            cur.execute(
                "UPDATE jobs SET status=%s, last_ranked_at=NOW() WHERE job_label=%s",
                (status, job_label)
            )
        else:
            cur.execute(
                "UPDATE jobs SET status=%s WHERE job_label=%s",
                (status, job_label)
            )
    conn.close()

# ── Candidate queries ──────────────────────────────────────────────────────────

def fetch_candidates(conn, job_label: str) -> pd.DataFrame:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT apply_id, candidate_name, email, mobile, location,
                   degree, university, experience_detail,
                   age, expected_salary, current_salary, application_date,
                   bdjobs_score, has_uploaded_cv, pdf_path,
                   overall_score, skills_score, experience_score,
                   leadership_score, education_score, culture_fit_score,
                   edu_tier_score, edu_degree_score, edu_gpa_score,
                   experience_years, strengths, gaps, risk_flags,
                   recommendation, reasoning, hr_override, hr_note,
                   ranked_at, rank_error, pdf_text_chars
            FROM candidates
            WHERE job_label = %s
            ORDER BY overall_score DESC NULLS LAST
        """, (job_label,))
        rows = cur.fetchall()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["experience_years"] = pd.to_numeric(df["experience_years"], errors="coerce")
    df["age"]              = pd.to_numeric(df["age"], errors="coerce")
    df["ranked_at"]        = pd.to_datetime(df["ranked_at"]).dt.strftime("%d %b %Y %H:%M")
    return df


def save_hr_override(job_label: str, apply_id: str, override: str, note: str,
                      hr_user: str = "HR"):
    """Update hr_override on the candidates row AND append a row to
    hr_audit_log (Part 7.5) so the change is traceable for compliance."""
    conn = fresh_conn()
    with conn.cursor() as cur:
        cur.execute(
            "SELECT hr_override FROM candidates WHERE job_label=%s AND apply_id=%s",
            (job_label, apply_id),
        )
        old_row = cur.fetchone()
        old_value = old_row[0] if old_row else None

        cur.execute(
            "UPDATE candidates SET hr_override=%s, hr_note=%s "
            "WHERE job_label=%s AND apply_id=%s",
            (override, note, job_label, apply_id),
        )
        cur.execute(
            """
            INSERT INTO hr_audit_log
                (job_label, apply_id, hr_user, old_value, new_value, note)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (job_label, apply_id, hr_user, old_value, override, note),
        )
    conn.close()


def fetch_audit_log(job_label: str | None = None,
                    apply_id:  str | None = None) -> pd.DataFrame:
    """Return the HR override audit log, optionally filtered by job or candidate."""
    conn = fresh_conn()
    where: list[str] = []
    params: list = []
    if job_label:
        where.append("job_label = %s"); params.append(job_label)
    if apply_id:
        where.append("apply_id = %s");  params.append(apply_id)
    clause = ("WHERE " + " AND ".join(where)) if where else ""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"""
            SELECT id, job_label, apply_id, hr_user, old_value, new_value,
                   note, changed_at
              FROM hr_audit_log
              {clause}
             ORDER BY changed_at DESC
             LIMIT 500
        """, params)
        rows = cur.fetchall()
    conn.close()
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def find_duplicate_applications(apply_id: str | None = None,
                                 email:    str | None = None,
                                 mobile:   str | None = None,
                                 exclude_job: str | None = None) -> list[dict]:
    """Part 7.7 — find other job postings the same person applied to.

    Match by email OR mobile when either is supplied (case-insensitive).
    """
    if not (email or mobile):
        return []
    conn = fresh_conn()
    params: list = []
    clauses: list[str] = []
    if email:
        clauses.append("LOWER(email) = LOWER(%s)"); params.append(email)
    if mobile:
        clauses.append("mobile = %s"); params.append(mobile)
    sql = f"""
        SELECT c.apply_id, c.candidate_name, c.job_label,
               j.job_title, c.recommendation, c.overall_score
          FROM candidates c
          LEFT JOIN jobs j ON j.job_label = c.job_label
         WHERE ({" OR ".join(clauses)})
    """
    if exclude_job:
        sql += " AND c.job_label <> %s"
        params.append(exclude_job)
    sql += " ORDER BY c.overall_score DESC NULLS LAST LIMIT 10"
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows] if rows else []

# ── Global stats ───────────────────────────────────────────────────────────────

# ── Excel export ───────────────────────────────────────────────────────────────

def to_excel(df: pd.DataFrame, job_label: str) -> bytes:
    export = df.copy()
    for col in ["strengths","gaps","risk_flags"]:
        if col in export.columns:
            export[col] = export[col].apply(
                lambda x: "; ".join(x) if isinstance(x, list) else str(x or "")
            )
    for col in ["_idx","pdf_text_chars","rank_error","hr_override","hr_note","has_uploaded_cv"]:
        if col in export.columns:
            export = export.drop(columns=[col])
    rename = {
        "apply_id":"Apply ID","candidate_name":"Candidate Name",
        "overall_score":"AI Overall Score","recommendation":"Recommendation",
        "skills_score":"Skills Score","experience_score":"Experience Score",
        "leadership_score":"Leadership Score","education_score":"Education Score",
        "edu_tier_score":"Edu · Tier","edu_degree_score":"Edu · Degree","edu_gpa_score":"Edu · GPA",
        "culture_fit_score":"Culture Fit Score","experience_years":"Experience (yrs)",
        "strengths":"Strengths","gaps":"Weaknesses","risk_flags":"Risk Flags",
        "reasoning":"AI Reasoning","bdjobs_score":"BDJobs Score (ref)",
        "application_date":"Application Date","age":"Age",
        "expected_salary":"Expected Salary","current_salary":"Current Salary",
        "pdf_path":"Resume File","email":"Email","mobile":"Mobile",
        "location":"Location","degree":"Degree","university":"University",
    }
    export = export.rename(columns={k:v for k,v in rename.items() if k in export.columns})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Rankings")
        ws = writer.sheets["Rankings"]
        from openpyxl.styles import Font, PatternFill, Alignment
        hfill = PatternFill("solid", fgColor="C8102E")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        rec_col = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Recommendation":
                rec_col = i
        rec_colors = {
            "Shortlist": ("DCFCE7","166534"),
            "Maybe":     ("FEF3C7","92400E"),
            "Reject":    ("FEE2E2","991B1B"),
        }
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            bg = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg)
            if rec_col:
                rc = ws.cell(row=row_idx, column=rec_col)
                if str(rc.value) in rec_colors:
                    rbg, rfg = rec_colors[str(rc.value)]
                    rc.fill = PatternFill("solid", fgColor=rbg)
                    rc.font = Font(bold=True, color=rfg)
        for col_cells in ws.columns:
            w = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w+3, 60)
        ws.row_dimensions[1].height = 22
    return buf.getvalue()

# ── Prompt preview ─────────────────────────────────────────────────────────────

def build_prompt_preview(job_data: dict) -> str:
    lines = ["=== AI SCORING PROMPT PREVIEW ===\n"]
    lines.append(f"ROLE: {job_data.get('job_title','Not specified')}")
    if job_data.get("department"):
        lines.append(f"DEPARTMENT: {job_data['department']}")
    if job_data.get("jd_text"):
        jd = job_data["jd_text"]
        lines.append(f"\nJOB DESCRIPTION:\n{jd[:800]}{'...' if len(jd)>800 else ''}")
    if job_data.get("required_skills"):
        lines.append(f"\nREQUIRED SKILLS: {', '.join(job_data['required_skills'])}")
    if job_data.get("min_experience") and job_data["min_experience"] != "Any":
        lines.append(f"MINIMUM EXPERIENCE: {job_data['min_experience']}")
    if job_data.get("education_req") and job_data["education_req"] != "Any":
        lines.append(f"EDUCATION: {job_data['education_req']}")
    ws = job_data.get("weight_skills",     50)
    we = job_data.get("weight_exp",        30)
    wu = job_data.get("weight_edu",        20)
    wl = job_data.get("weight_leadership", 10)
    wc = job_data.get("weight_culture",     5)
    lines.append(
        f"\nSCORING WEIGHTS: Skills {ws}% / Experience {we}% / "
        f"Education {wu}% / Leadership {wl}% / Culture Fit {wc}%"
    )
    if job_data.get("red_flags"):
        lines.append("\nRED FLAGS:\n" + "\n".join(f"  - {f}" for f in job_data["red_flags"]))
    if job_data.get("interviewer_notes"):
        lines.append(f"\nADDITIONAL NOTES:\n{job_data['interviewer_notes']}")
    lines.append("""\n\nOUTPUT FORMAT (strict JSON):
{
  "overall_score": 0-100,
  "skills_score": 0-100,
  "experience_score": 0-100,
  "leadership_score": 0-100,
  "education_score": 0-100,
  "culture_fit_score": 0-100,
  "experience_years": number,
  "strengths": ["..."],
  "gaps": ["..."],
  "recommendation": "Shortlist|Maybe|Reject",
  "reasoning": "2-sentence summary"
}""")
    return "\n".join(lines)

# ── Theme ──────────────────────────────────────────────────────────────────────

def init_theme():
    st.session_state["day_mode"] = True

def theme_toggle():
    st.session_state["day_mode"] = True

def get_css() -> str:
    return CSS

def t(day_val, night_val=None):
    return day_val

# ── Sidebar ────────────────────────────────────────────────────────────────────

def render_sidebar():
    with st.sidebar:
        logo_path = LOGO_PATH
        if os.path.exists(logo_path):
            st.markdown('<div style="padding:2% 2% 1rem 2%;">', unsafe_allow_html=True)
            st.image(logo_path, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown("""
                <div style="padding:0.6rem 0 1.4rem;">
                    <div style="font-size:1.05rem;font-weight:700;color:#FFFFFF !important;">Olympic Industries PLC</div>
                    <div style="font-size:0.72rem;color:rgba(255,255,255,0.7) !important;margin-top:3px;text-transform:uppercase;letter-spacing:0.1em;">HR Intelligence Platform</div>
                </div>
            """, unsafe_allow_html=True)

        st.markdown('<div class="nav-label">Navigation</div>', unsafe_allow_html=True)
        st.page_link("Home.py",                           label="📋  Dashboard")
        st.page_link("pages/1_Department_Rankings.py",    label="🏢  Department Rankings")
        st.page_link("pages/2_Job_Rankings.py",           label="📊  Job Rankings")
        st.page_link("pages/3_New_Job.py",                label="📝  New Job Posting")
        st.page_link("pages/4_Processing_Status.py",      label="⏳  Processing Status")
        st.page_link("pages/5_Settings.py",               label="⚙️  Settings")

def fetch_global_stats(conn=None) -> dict:
    """Global candidate stats. `conn` is accepted for API symmetry but we
    always open a fresh autocommit connection so Streamlit cache can't serve
    a stale snapshot."""
    fresh = fresh_conn()
    with fresh.cursor() as cur:
        cur.execute("""
            SELECT
                COUNT(DISTINCT job_label)                               AS total_jobs,
                COUNT(*)                                                AS total_candidates,
                SUM(CASE WHEN overall_score IS NULL THEN 1 ELSE 0 END) AS pending,
                ROUND(AVG(overall_score))                               AS avg_score
            FROM candidates
        """)
        row = cur.fetchone()
    fresh.close()
    return {
        "total_jobs":       row[0] or 0,
        "total_candidates": row[1] or 0,
        "pending":          row[2] or 0,
        "avg_score":        row[3] or 0,
    }

# ── CSS ────────────────────────────────────────────────────────────────────────

CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

#MainMenu, footer, header { visibility: hidden !important; }
[data-testid="collapsedControl"]        { display: none !important; }
[data-testid="stSidebarCollapseButton"] { display: none !important; }
[data-testid="stSidebar"] {
    min-width: 288px !important;
    max-width: 288px !important;
    transform: none !important;
}
[data-testid="stSidebar"][aria-expanded="false"] {
    min-width: 288px !important;
    transform: none !important;
    margin-left: 0 !important;
}

[data-testid="stSidebarNav"] { display: none !important; }

.block-container { padding: 1.8rem 2.2rem 3rem !important; max-width: 1440px; }

html, body, p, span, div, label, li, td, th, input, textarea, select {
    font-family: 'Inter', sans-serif !important;
    font-size: 15px;
}

[data-testid="stSidebar"] { background-color: #C8102E !important; border-right: none !important; }
[data-testid="stSidebar"] > div { background-color: #C8102E !important; }
[data-testid="stSidebar"] * { color: #FFFFFF !important; font-family: 'Inter', sans-serif !important; font-size: 14px !important; }
[data-testid="stSidebar"] a { color: #FFFFFF !important; text-decoration: none !important; }
[data-testid="stSidebar"] a:hover { color: rgba(255,255,255,0.75) !important; }
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.12) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    color: #FFFFFF !important;
    font-size: 0.83rem !important;
    width: 100% !important;
    border-radius: 6px !important;
    padding: 0.45rem 0.75rem !important;
    margin-bottom: 3px !important;
    text-align: left !important;
    font-weight: 400 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.22) !important;
    border-color: rgba(255,255,255,0.4) !important;
}

.nav-label {
    font-size: 0.68rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: rgba(255,255,255,0.6) !important;
    margin-bottom: 0.5rem;
    padding-bottom: 0.3rem;
    border-bottom: 1px solid rgba(255,255,255,0.2);
}

[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background-color: rgba(255,255,255,0.2) !important;
    border-radius: 4px !important;
}
[data-testid="stSidebar"] span[data-baseweb="tag"] * {
    color: #FFFFFF !important;
}

html, body { background-color: #F8FAFC !important; }
.stApp, section[data-testid="stMain"], section[data-testid="stMain"] > div,
.block-container { background-color: #F8FAFC !important; }

html, body, p, span, label, li, td, th { color: #1E293B !important; }

.page-title { font-size: 1.55rem; font-weight: 700; line-height: 1.2; margin-bottom: 0; color: #1E293B !important; }
.page-sub   { font-size: 0.85rem; margin-top: 3px; color: #64748B !important; }
.section-hd {
    font-size: 0.72rem; font-weight: 600; letter-spacing: 0.14em;
    text-transform: uppercase; margin-bottom: 0.75rem;
    padding-bottom: 0.45rem; color: #64748B !important;
    border-bottom: 1px solid #E2E8F0;
}
.divider { border: none; border-top: 1px solid #E2E8F0; margin: 1.2rem 0; }

.metric-card {
    border-radius: 10px; padding: 1.1rem 1.3rem;
    background: #FFFFFF; border: 1px solid #E2E8F0;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.metric-val { font-size: 2.1rem; font-weight: 700; line-height: 1; color: #1E293B !important; }
.metric-lbl { font-size: 0.74rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 4px; color: #64748B !important; }

.verdict-badge     { display: inline-block; padding: 3px 11px; border-radius: 12px; font-size: 0.78rem; font-weight: 600; }
.verdict-shortlist { background: #DCFCE7; color: #166534 !important; }
.verdict-maybe     { background: #FEF3C7; color: #92400E !important; }
.verdict-reject    { background: #FEE2E2; color: #991B1B !important; }

.status-badge      { display: inline-block; padding: 2px 9px; border-radius: 10px; font-size: 0.73rem; font-weight: 500; }
.status-complete   { background: #DCFCE7; color: #166534 !important; }
.status-processing { background: #DBEAFE; color: #1E40AF !important; }
.status-pending    { background: #F1F5F9; color: #475569 !important; }
.status-error      { background: #FEE2E2; color: #991B1B !important; }

.score-pill   { display: inline-block; border-radius: 6px; padding: 4px 12px; font-size: 0.8rem; font-weight: 600; margin: 3px 4px 3px 0; }
.score-green  { background: #DCFCE7; color: #166534 !important; }
.score-yellow { background: #FEF3C7; color: #92400E !important; }
.score-red    { background: #FEE2E2; color: #991B1B !important; }

.flag-chip { display: inline-block; border-radius: 4px; padding: 2px 8px; font-size: 0.73rem; font-weight: 500; margin: 2px 3px 2px 0; background: #FEE2E2; color: #991B1B !important; }

.cand-name-lg { font-size: 1.4rem; font-weight: 700; line-height: 1.2; }
.cand-meta-sm { font-size: 0.83rem; margin-top: 3px; }
.hint-text    { font-size: 0.8rem; color: #94A3B8 !important; }

div[data-testid="stDataFrame"] { border: 1px solid #E2E8F0 !important; border-radius: 8px !important; box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important; }
div[data-testid="stDataFrame"] thead th {
    background-color: #C8102E !important;
    color: #FFFFFF !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.03em !important;
}
div[data-testid="stDataFrame"] { color-scheme: light !important; }

div[data-baseweb="select"] > div { background-color: #FFFFFF !important; border-color: #CBD5E1 !important; border-radius: 6px !important; }
div[data-baseweb="select"] * { color: #1E293B !important; }
div[data-baseweb="popover"] { background-color: #FFFFFF !important; border: 1px solid #E2E8F0 !important; border-radius: 8px !important; }
div[data-baseweb="popover"] * { background-color: #FFFFFF !important; color: #1E293B !important; }

# _________Sidebar Tag Overrides (for better visibility against red background)_________
[data-testid="stSidebar"] span[data-baseweb="tag"] { background-color: rgba(0,0,0,0.25) !important; border-radius: 4px !important; }
[data-testid="stSidebar"] span[data-baseweb="tag"] * { color: #6E1400 !important; }

span[data-baseweb="tag"] { background-color: #FECDD3 !important; border-radius: 4px !important; }
span[data-baseweb="tag"] * { color: #9F1239 !important; }

textarea, input[type="text"] { background-color: #FFFFFF !important; border: 1px solid #CBD5E1 !important; border-radius: 6px !important; color: #1E293B !important; }

button[kind="primary"], button[data-testid="baseButton-primary"] { background: #C8102E !important; color: #FFFFFF !important; border: none !important; border-radius: 6px !important; font-family: 'Inter', sans-serif !important; font-size: 0.85rem !important; font-weight: 500 !important; }
button[kind="primary"] p, button[data-testid="baseButton-primary"] p { color: #FFFFFF !important; }
button[kind="primary"] span, button[data-testid="baseButton-primary"] span { color: #FFFFFF !important; }
button[kind="primary"]:hover, button[data-testid="baseButton-primary"]:hover { background: #A00D24 !important; }
button[kind="secondary"], button[data-testid="baseButton-secondary"] { background: #FFFFFF !important; color: #C8102E !important; border: 1px solid #E2E8F0 !important; border-radius: 6px !important; font-family: 'Inter', sans-serif !important; font-size: 0.85rem !important; font-weight: 500 !important; }
button[kind="secondary"]:hover, button[data-testid="baseButton-secondary"]:hover { background: #FFF1F2 !important; border-color: #C8102E !important; color: #A00D24 !important; }
.stDownloadButton > button { background: #FFFFFF !important; border: 1px solid #CBD5E1 !important; color: #C8102E !important; border-radius: 6px !important; font-family: 'Inter', sans-serif !important; font-size: 0.85rem !important; font-weight: 500 !important; }
.stDownloadButton > button:hover { border-color: #C8102E !important; background: #FFF1F2 !important; }

div[data-testid="stAlert"] { border-radius: 8px !important; font-size: 0.88rem !important; }
.stMarkdown p, .stMarkdown li { color: #374151 !important; font-size: 0.92rem; line-height: 1.65; }

/* ── Department cards (Department Rankings page + Home grouped view) ── */
.dept-card {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-left: 4px solid #C8102E;
    border-radius: 10px;
    padding: 1.2rem 1.4rem;
    margin-bottom: 0.7rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.dept-name {
    font-size: 1.05rem;
    font-weight: 700;
    color: #1E293B !important;
    margin-bottom: 6px;
    letter-spacing: 0.01em;
}
.dept-stats {
    font-size: 0.85rem;
    color: #475569 !important;
}
.dept-meta {
    font-size: 0.72rem;
    color: #94A3B8 !important;
    margin-top: 4px;
    letter-spacing: 0.02em;
}
.dept-group-title {
    font-size: 0.78rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #C8102E !important;
    padding: 0.8rem 0 0.3rem;
    border-bottom: 1px solid #E2E8F0;
    margin-bottom: 0.5rem;
}
.rank-pill {
    display: inline-block;
    background: #C8102E;
    color: #FFFFFF !important;
    border-radius: 10px;
    padding: 2px 10px;
    font-size: 0.76rem;
    font-weight: 700;
    letter-spacing: 0.02em;
}

/* Preserve Material Symbols icon font everywhere (main + sidebar). The
   global `span` and `[data-testid="stSidebar"] *` font-family overrides
   above would otherwise render expander chevrons as literal ligature
   text such as "keyboard_arrow_down". Placed last so cascade wins. */
[data-testid="stIconMaterial"],
[data-testid="stIconMaterial"] *,
[data-testid="stSidebar"] [data-testid="stIconMaterial"],
[data-testid="stSidebar"] [data-testid="stIconMaterial"] * {
    font-family: 'Material Symbols Rounded', 'Material Icons' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}
</style>"""

CUSTOM_CSS = CSS