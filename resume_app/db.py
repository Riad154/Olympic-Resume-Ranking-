"""
db.py — Data layer, theming, and shared utilities.
Olympic Industries PLC — HR Intelligence Platform
"""

import io
import os
import re
import csv
import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).resolve().parent.parent / ".env"
    if _env_path.exists():
        load_dotenv(str(_env_path), override=True)
except Exception:
    pass

import psycopg2
import psycopg2.extras
import pandas as pd
import streamlit as st

# ── BDJobs Registry (43 live listings) ─────────────────────────────────────────
try:
    from resume_app._bdjobs_registry import BDJOBS_JOB_REGISTRY
except ImportError:
    from _bdjobs_registry import BDJOBS_JOB_REGISTRY

# ── Favicon (Olympic logo) ────────────────────────────────────────────────────
try:
    from _favicon import OLYMPIC_FAVICON as FAVICON
except Exception:
    FAVICON = "📋"

# ── Config ─────────────────────────────────────────────────────────────────────

def _pg_conf():
    """Build PG connection dict from env vars (local/GitHub Actions) or
    st.secrets (Streamlit Cloud). Returns empty dict if not configured."""
    # Read env vars FIRST — st.secrets.get() has a side-effect that
    # can clear os.environ when run outside a Streamlit app context.
    env_host = (os.environ.get("PG_HOST", "") or "").strip()
    env_user = (os.environ.get("PG_USER", "") or "").strip()
    env_password = (os.environ.get("PG_PASSWORD", "") or "").strip()

    # Try st.secrets first (Streamlit Cloud)
    try:
        pg = st.secrets.get("postgresql", {})
        if pg:
            host = (pg.get("host", "") or "").strip()
            user = (pg.get("user", "") or "").strip()
            password = (pg.get("password", "") or "").strip()
            if not host or not user:
                return {}  # Not configured
            port_str = str(pg.get("port", "5432") or "5432")
            return {
                "host":     host,
                "port":     int(port_str) if port_str else 5432,
                "dbname":   (pg.get("dbname", "") or "").strip() or "resume_ranking",
                "user":     user,
                "password": password,
            }
    except Exception:
        pass

    # Fall back to environment variables (local / GitHub Actions)
    if not env_host or not env_user:
        return {}  # Not configured
    port_str = os.environ.get("PG_PORT", "5432") or "5432"
    return {
        "host":     env_host,
        "port":     int(port_str),
        "dbname":   (os.environ.get("PG_DBNAME", "") or "").strip() or "resume_ranking",
        "user":     env_user,
        "password": env_password,
    }

PG_CONN = _pg_conf()

def pg_is_configured() -> bool:
    """Return True if PostgreSQL credentials are set."""
    return bool(PG_CONN.get("host") and PG_CONN.get("user"))

# Project root = parent of resume_app/
PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESUMES_BASE = os.environ.get("RESUMES_BASE", str(PROJECT_ROOT / "downloaded_resumes"))
RANKER_PATH  = os.environ.get("RANKER_PATH",  str(PROJECT_ROOT / "ranker.py"))
VENV_PYTHON  = os.environ.get("VENV_PYTHON",  sys.executable)
LOGO_PATH    = str(Path(__file__).resolve().parent / "plc_logo_w_text.png")

DEPARTMENTS = [
    "Finance and Accounts",
    "Admin",
    "AI & Digital Transformation",
    "Brand & Marketing",
    "Corporate Affairs",
    "Customer Service Department (CSD)",
    "Delivery",
    "Distribution",
    "Engineering",
    "Export",
    "External Audit",
    "Factory Administration",
    "Field Force",
    "Human Resource (HR)",
    "Information & Communication Technology (ICT)",
    "Import",
    "Institutional Sales",
    "Internal Audit",
    "Legal Affairs",
    "Local Procurement",
    "Management",
    "Market Audit",
    "Mechanical",
    "Management Information System (MIS)",
    "Operations",
    "Plastic Production",
    "Supply Chain",
    "Production",
    "Quality Assurance Department (QAD)",
    "ERP - SAP",
    "Sales",
    "Secretariat",
    "Security",
    "Share",
    "Store",
    "Transport",
    "VAT / VAT & Delivery",
]

# Authoritative department list used by ranker --department, New Job form,
# assign-dept UI, and nav filters.  "Uncategorized" is the first-class
# fallback for legacy jobs with no department assigned.
DEPARTMENT_LIST = ["Uncategorized"] + DEPARTMENTS

EXPERIENCE_OPTIONS = ["Any","1 year","2 years","3 years","5 years","7 years","10 years","15+ years"]
EDUCATION_OPTIONS  = ["Any","SSC","HSC","Diploma","Bachelor's","Master's","MBA","PhD","Professional Certification"]

COMMON_SKILLS = [
    # Leadership & Management
    "Team Leadership","People Management","Strategic Planning","Change Management",
    "Performance Management","Stakeholder Management","Budget Management",
    "Cross-functional Collaboration","Conflict Resolution","Decision Making",

    # Finance & Accounts
    "Financial Reporting","Budgeting & Forecasting","Cost Control",
    "Accounts Payable","Accounts Receivable","Tax Compliance",
    "Audit","Payroll Management","Treasury Management","VAT Compliance",
    "Tally","QuickBooks",

    # Sales & Marketing
    "Sales Management","Key Account Management","Trade Marketing",
    "Brand Management","Market Research","Digital Marketing",
    "Retail Management","Distribution Management","Field Sales",
    "Institutional Sales","Territory Management","Trade Promotion",

    # Supply Chain & Logistics
    "Supply Chain Management","Logistics","Procurement","Inventory Management",
    "Production Planning","Demand Forecasting","Warehouse Management",
    "Import & Export","Customs Clearance","Fleet Management","Cold Chain",

    # HR
    "Recruitment","Training & Development","HR Policy","Labor Law",
    "Employee Relations","Compensation & Benefits",
    "Organizational Development","Talent Management","HRIS",

    # Production & Quality
    "Production Management","Quality Control","Quality Assurance",
    "Six Sigma","Lean Manufacturing","GMP","HACCP","ISO Standards",
    "Process Improvement","FMCG Manufacturing","Food Safety",

    # Engineering & Maintenance
    "Mechanical Engineering","Electrical Engineering","Preventive Maintenance",
    "Equipment Troubleshooting","AutoCAD","PLC","Industrial Automation",

    # IT & Systems
    "SAP","ERP","Power BI","Excel","SQL","Data Analysis",
    "Process Automation","Python","Machine Learning","Network Administration",
    "IT Support","MIS Reporting","Database Management",

    # Admin & General
    "MS Office","Report Writing","Presentation Skills","Negotiation",
    "Vendor Management","Contract Management","Compliance","Documentation",
    "FMCG Experience","Project Management","Agile",
]


SKILL_DOMAINS = {
    "🤖 AI & Technology": [
        "Machine Learning","Deep Learning","Computer Vision","NLP","Python",
        "Data Analysis","Power BI","Tableau","SQL","Process Automation",
        "RPA","FastAPI","Docker","n8n","ERP Implementation","Digital Strategy",
        "AI Project Management","MIS Reporting","Report Automation","ETL",
        "Database Management","Dashboard Development",
    ],
    "💰 Finance & Audit": [
        "Financial Reporting","Budgeting & Forecasting","Cost Control","Cost Accounting",
        "Accounts Payable","Accounts Receivable","Tax Compliance","VAT Compliance",
        "Audit","Payroll Management","Treasury Management","Cash Flow Management",
        "Bank Reconciliation","Fixed Assets Management","Tally","SAP FICO",
        "Internal Audit","External Audit","Risk Assessment","Financial Controls",
        "Process Audit","Stock Audit","Compliance Audit",
    ],
    "📦 Supply Chain & Logistics": [
        "Supply Chain Management","Inventory Management","Demand Planning",
        "Warehouse Management","S&OP","Logistics Coordination","Stock Reconciliation",
        "Cold Chain Management","3PL Management","Supply Chain Analytics",
        "Fleet Management","Route Planning","Vehicle Maintenance","Driver Management",
        "Delivery Scheduling","Last Mile Delivery","Store Management",
        "FIFO/FEFO Management","Material Handling","GRN Processing","Stock Audit",
    ],
    "🏭 Production & Quality": [
        "Production Planning","Production Management","OEE Improvement",
        "Line Balancing","Shift Management","FMCG Manufacturing",
        "Injection Moulding","Blow Moulding","Plastic Processing",
        "Machine Operation","Downtime Reduction","Batch Production",
        "Quality Control","Quality Assurance","GMP","HACCP","ISO 22000",
        "FSSC 22000","Food Safety","In-process Quality Check",
        "Raw Material Inspection","Finished Goods Testing",
        "Microbiology","Laboratory Management","SOP Development",
    ],
    "⚙️ Engineering & Maintenance": [
        "Mechanical Engineering","Electrical Engineering","Preventive Maintenance",
        "Breakdown Maintenance","Equipment Troubleshooting","AutoCAD",
        "PLC Programming","Pneumatics & Hydraulics","Boiler Operation",
        "Utility Management","Energy Management","Fabrication","Welding",
    ],
    "📊 Sales & Marketing": [
        "Sales Management","Key Account Management","Trade Marketing",
        "Brand Management","Market Research","Digital Marketing",
        "Retail Management","Distribution Management","Field Sales",
        "Institutional Sales","Territory Management","Trade Promotion",
        "Campaign Management","Consumer Insights","Media Planning",
        "ATL/BTL Marketing","Social Media Marketing","Content Creation",
        "Marketing Analytics","New Product Development","Packaging Development",
        "Beat Planning","Outlet Expansion","Market Development",
    ],
    "🚚 Distribution & Field": [
        "Distribution Network Management","Channel Management","Distributor Management",
        "Secondary Sales","Route to Market","Coverage Expansion",
        "Van Sales","Depot Management","Stock Replenishment",
        "Market Survey","Competitor Analysis","Price Monitoring",
        "Outlet Audit","Shelf Share Analysis","Field Data Collection",
        "Market Intelligence","Distribution Audit",
    ],
    "👥 Human Resource": [
        "Recruitment & Selection","Training & Development","HR Policy Development",
        "Performance Management","Compensation & Benefits","Employee Relations",
        "Organizational Development","Talent Management","HRIS","Succession Planning",
        "HR Compliance","Labor Law","Grievance Handling","Onboarding",
    ],
    "🖥️ ICT & MIS": [
        "Network Administration","IT Infrastructure","Cybersecurity",
        "ERP Support","Hardware & Software Troubleshooting","Server Management",
        "Active Directory","IT Helpdesk","Cloud Computing","Backup & Recovery",
        "IT Procurement","CCTV & Access Control","Advanced Excel",
        "KPI Reporting","Python",
    ],
    "🔵 SAP & ERP": [
        "SAP MM","SAP SD","SAP FI","SAP CO","SAP PP","SAP QM","SAP WM",
        "SAP HR","SAP BASIS","SAP ABAP","S/4HANA","SAP Implementation",
        "SAP Configuration","SAP Reporting","SAP Integration",
    ],
    "🌐 Export & Import": [
        "Export Documentation","LC Management","Customs Clearance",
        "BGMEA/BEPZA Compliance","Freight Forwarding","Incoterms",
        "Shipping & Logistics","Import Documentation","Bond Licensing",
        "Foreign Currency Management","Trade Finance",
    ],
    "📋 Admin & Compliance": [
        "Office Management","Facility Management","Vendor Management",
        "Contract Management","Record Keeping","Travel & Logistics Coordination",
        "Event Management","Document Control","Scheduling & Calendar Management",
        "Corporate Governance","Regulatory Compliance","Legal Documentation",
        "Contract Drafting & Review","Company Secretarial Work","RJSC Compliance",
        "BIDA Compliance","Intellectual Property","Due Diligence",
        "Factory Compliance","Industrial Relations","Factory Act Compliance",
    ],
    "🛡️ Security & Store": [
        "Physical Security","Access Control","CCTV Monitoring",
        "Security Risk Assessment","Crisis Management","Guard Management",
        "Loss Prevention","Inventory Control","Bin Card Management",
        "Warehouse Layout Optimization",
    ],
    "🏢 Leadership & Strategy": [
        "Team Leadership","People Management","Strategic Planning","Change Management",
        "Performance Management","Stakeholder Management","Budget Management",
        "Cross-functional Collaboration","Conflict Resolution","Decision Making",
        "Operations Management","Project Management","Business Process Improvement",
        "KPI Management","Agile","Six Sigma",
    ],
    "📝 General & Communication": [
        "Report Writing","Presentation Skills","Negotiation","Problem Solving",
        "Data Entry","Filing & Documentation","FMCG Experience",
        "Communication Skills","Team Coordination","MS Office",
    ],
}

# ── Department Skill Profiles (Part 2) ─────────────────────────────────────────
# Used by ranker.py to inject department-specific scoring context into the prompt.
# Each entry defines:
#   "core_skills"     : must-have tools/skills. Missing >50% caps skills_score at 50.
#   "bonus_skills"    : differentiators that push score above 75.
#   "anti_skills"     : skills that are irrelevant and should not inflate score.
#   "scoring_note"    : free-text guidance injected directly into LLM prompt.

DEPARTMENT_SKILL_PROFILES = {

    "Brand & Marketing": {
        "core_skills": [
            # Brand management core
            "Brand Strategy", "Campaign Management", "New Product Development",
            "Market Research", "Consumer Insights", "Marketing Analytics",
            "ATL/BTL Marketing", "Social Media Marketing",
            # Designer core (for Graphic Designer / Media Coordinator roles)
            "Adobe Premiere Pro", "Adobe Photoshop", "Adobe Illustrator",
            "Adobe InDesign", "Canva", "Video Editing",
            # Packaging (for Graphic Designer)
            "Packaging Design", "Print-Ready File Preparation",
        ],
        "bonus_skills": [
            "CapCut", "After Effects", "Motion Graphics",
            "Google Analytics", "Meta Ads Manager", "Marketing Analytics",
            "New Product Development", "Trade Marketing",
            "Agency Coordination", "OVC Production Coordination",
            "AI Creative Tools",
        ],
        "anti_skills": ["PLC Programming", "Boiler Operation", "Tax Compliance", "Audit"],
        "scoring_note": (
            "Brand & Marketing at Olympic has two distinct sub-categories:\n"
            "  BRAND ROLES (Executive/Sr. Executive/AM - Brand): Weight market analysis, "
            "NPD, campaign strategy, and FMCG food experience. BBA/MBA Marketing preferred.\n"
            "  DESIGN/MEDIA ROLES (Graphic Designer, Media Coordinator): Weight creative tools "
            "heavily. Adobe Creative Suite (Illustrator + Photoshop + InDesign) are mandatory "
            "for Graphic Designer. Adobe Premiere Pro + CapCut mandatory for Media Coordinator. "
            "Portfolio evidence is more important than degree for design roles.\n"
            "  PACKAGING DESIGN: FMCG packaging portfolio is the primary differentiator. "
            "Candidates without packaging design experience cap at 55 on skills_score."
        ),
    },

    "AI & Digital Transformation": {
        "core_skills": [
            "Python", "Machine Learning", "Data Analysis", "SQL",
            "Power BI", "ERP Implementation", "Process Automation",
            "Digital Strategy", "AI Project Management", "Dashboard Development",
        ],
        "bonus_skills": [
            "Deep Learning", "NLP", "Computer Vision", "Docker",
            "FastAPI", "n8n", "RPA", "ETL", "Tableau", "MIS Reporting",
            "S/4HANA", "SAP Integration", "Change Management",
            "LLM Fine-tuning", "Vector Databases", "RAG Pipelines",
            "Prompt Engineering", "MLOps", "Azure/AWS/GCP",
        ],
        "anti_skills": ["Manual Data Entry", "Physical Security", "Boiler Operation"],
        "scoring_note": (
            "For AI & Digital Transformation roles, differentiate by seniority:\n"
            "  SENIOR EXECUTIVE / MANAGER+: Must show at least one end-to-end AI/automation "
            "project deployed in a production environment (not just academic). "
            "ERP (SAP) implementation experience is highly valued in FMCG context. "
            "Digital strategy ownership and measurable ROI from DT initiatives are required.\n"
            "  EXECUTIVE / OFFICER: Python + SQL proficiency with a deployed data pipeline or "
            "dashboard is sufficient. Theoretical-only candidates (no deployed project) cap at 65.\n"
            "  GENERAL RULE: Candidates with only theoretical knowledge and no production "
            "deployments should score no higher than 65 on skills_score regardless of qualifications."
        ),
    },

    "Finance and Accounts": {
        "core_skills": [
            "Financial Reporting", "Budgeting & Forecasting", "Cost Control",
            "Accounts Payable", "Accounts Receivable", "Tax Compliance",
            "VAT Compliance", "Tally", "SAP FICO", "Bank Reconciliation",
            "Fixed Assets Management", "Payroll Management",
        ],
        "bonus_skills": [
            "Treasury Management", "Cash Flow Management", "Cost Accounting",
            "Internal Audit", "External Audit", "Risk Assessment",
            "Financial Controls", "IFRS", "Bangladesh Tax Law", "NBFI Regulations",
        ],
        "anti_skills": ["Video Editing", "CAD", "Machine Operation"],
        "scoring_note": (
            "For Finance roles: CA / ACCA / CPA qualification is a strong differentiator "
            "and should push education_score upward. VAT Compliance and Bangladesh "
            "Tax Law knowledge are essential for senior roles. SAP FICO is a bonus "
            "given Olympic's SAP environment."
        ),
    },

    "Human Resource (HR)": {
        "core_skills": [
            "Recruitment & Selection", "Training & Development", "HR Policy Development",
            "Performance Management", "Compensation & Benefits", "Labor Law",
            "Employee Relations", "HRIS", "Onboarding", "HR Compliance",
        ],
        "bonus_skills": [
            "Organizational Development", "Talent Management", "Succession Planning",
            "Grievance Handling", "Bangladesh Labor Law", "SAP HR",
            "Learning Management System (LMS)", "Employer Branding",
        ],
        "anti_skills": ["Mechanical Engineering", "PLC Programming", "Tax Compliance"],
        "scoring_note": (
            "For HR roles: knowledge of Bangladesh Labour Act 2006 is mandatory "
            "for mid/senior roles. Candidates with PGDHRM or CIPD qualification "
            "should receive a bonus on education_score. Factory HR experience "
            "is a differentiator in the FMCG manufacturing context."
        ),
    },

    "Information & Communication Technology (ICT)": {
        "core_skills": [
            "Network Administration", "IT Infrastructure", "Server Management",
            "Hardware & Software Troubleshooting", "Active Directory",
            "IT Helpdesk", "ERP Support", "Backup & Recovery",
        ],
        "bonus_skills": [
            "Cybersecurity", "Cloud Computing", "CCTV & Access Control",
            "Firewall Management", "VMware / Hyper-V", "IT Procurement",
            "SAP BASIS", "Python", "Advanced Excel",
        ],
        "anti_skills": ["Video Editing", "Brand Management", "Quality Control"],
        "scoring_note": (
            "For ICT roles: CCNA, CompTIA, or Microsoft certifications are strong "
            "education differentiators. Candidates supporting a large enterprise "
            "(500+ users) score higher on experience_score."
        ),
    },

    "ERP - SAP": {
        "core_skills": [
            "SAP MM", "SAP SD", "SAP FI", "SAP CO", "SAP PP",
            "SAP Configuration", "SAP Reporting", "S/4HANA",
            "SAP Implementation", "SAP Integration",
        ],
        "bonus_skills": [
            "SAP QM", "SAP WM", "SAP HR", "SAP ABAP", "SAP BASIS",
            "SAP Fiori", "SAP BW/BI", "End-User Training",
        ],
        "anti_skills": ["Manual Bookkeeping", "Physical Security", "Machine Operation"],
        "scoring_note": (
            "For SAP roles: at least 2 full-cycle SAP implementations are expected "
            "at the senior level. Module depth matters more than breadth. "
            "S/4HANA experience is a strong differentiator for Olympic's current roadmap."
        ),
    },

    "Supply Chain": {
        "core_skills": [
            "Supply Chain Management", "Inventory Management", "Demand Planning",
            "Warehouse Management", "S&OP", "Procurement", "Vendor Management",
            "Import & Export", "Logistics Coordination", "Stock Reconciliation",
        ],
        "bonus_skills": [
            "Cold Chain Management", "3PL Management", "Supply Chain Analytics",
            "FIFO/FEFO Management", "SAP MM", "Customs Clearance",
            "Demand Forecasting", "Cost Reduction Initiatives",
        ],
        "anti_skills": ["Video Editing", "HR Policy", "Tax Compliance"],
        "scoring_note": (
            "FMCG supply chain experience is a strong differentiator. "
            "Candidates who have managed high-SKU environments (200+ SKUs) "
            "or nationwide distribution networks score higher on experience_score."
        ),
    },

    "Sales": {
        "core_skills": [
            "Sales Management", "Key Account Management", "Trade Marketing",
            "Retail Management", "Distribution Management", "Field Sales",
            "Territory Management", "Beat Planning", "Outlet Expansion",
        ],
        "bonus_skills": [
            "Institutional Sales", "Trade Promotion", "Market Development",
            "Secondary Sales", "Van Sales", "Distributor Management",
            "CRM", "Sales Analytics", "Channel Management",
        ],
        "anti_skills": ["Software Development", "Mechanical Engineering", "Audit"],
        "scoring_note": (
            "For Sales roles: quantified achievements (revenue growth %, territory "
            "expansion, new outlet counts) are strong positive evidence and should "
            "increase experience_score. FMCG field sales experience is weighted heavily."
        ),
    },

    "Production": {
        "core_skills": [
            "Production Management", "Production Planning", "FMCG Manufacturing",
            "Shift Management", "OEE Improvement", "Line Balancing",
            "Batch Production", "Downtime Reduction", "Machine Operation",
        ],
        "bonus_skills": [
            "Lean Manufacturing", "Six Sigma", "Injection Moulding",
            "Blow Moulding", "Plastic Processing", "SOP Development",
            "5S Methodology", "TPM",
        ],
        "anti_skills": ["Digital Marketing", "Tax Compliance", "Network Administration"],
        "scoring_note": (
            "For Production roles: experience in high-volume FMCG or plastic "
            "manufacturing lines is essential. Lean / Six Sigma certification "
            "should push education_score upward."
        ),
    },

    "Quality Assurance Department (QAD)": {
        "core_skills": [
            "Quality Control", "Quality Assurance", "GMP", "HACCP",
            "ISO 22000", "Food Safety", "In-process Quality Check",
            "Raw Material Inspection", "Finished Goods Testing",
            "Laboratory Management", "SOP Development",
        ],
        "bonus_skills": [
            "FSSC 22000", "Microbiology", "Halal Certification",
            "ISO 9001", "ISO 14001", "Six Sigma", "Root Cause Analysis",
            "Sensory Evaluation", "Statistical Process Control (SPC)",
        ],
        "anti_skills": ["Software Development", "HR Policy", "Finance Reporting"],
        "scoring_note": (
            "For QAD roles: BSc/MSc in Food Science, Chemistry, Microbiology, "
            "or Chemical Engineering is ideal. HACCP Lead Auditor certification "
            "is a strong differentiator. Practical lab experience in FMCG is required."
        ),
    },

    "Engineering": {
        "core_skills": [
            "Mechanical Engineering", "Electrical Engineering",
            "Preventive Maintenance", "Breakdown Maintenance",
            "Equipment Troubleshooting", "PLC Programming",
            "Pneumatics & Hydraulics", "Utility Management",
        ],
        "bonus_skills": [
            "AutoCAD", "Boiler Operation", "Energy Management",
            "Industrial Automation", "Fabrication", "Welding",
            "HVAC", "Instrumentation",
        ],
        "anti_skills": ["Digital Marketing", "HR Recruitment", "Tax Filing"],
        "scoring_note": (
            "For Engineering/Maintenance roles: practical hands-on experience "
            "with industrial machinery is mandatory. BUET or equivalent engineering "
            "degree is highly valued. Candidates with no factory floor experience "
            "should score no higher than 55 on experience_score."
        ),
    },

    "Distribution": {
        "core_skills": [
            "Distribution Network Management", "Channel Management",
            "Distributor Management", "Secondary Sales", "Route to Market",
            "Coverage Expansion", "Depot Management", "Stock Replenishment",
        ],
        "bonus_skills": [
            "Van Sales", "Fleet Management", "Route Planning",
            "Market Survey", "Outlet Audit", "Distribution Audit",
            "Market Intelligence", "Field Data Collection",
        ],
        "anti_skills": ["Software Engineering", "Audit", "Legal"],
        "scoring_note": (
            "Distribution roles require deep field knowledge of Bangladesh's "
            "district and upazilla-level trade channels. National coverage "
            "management experience is a strong differentiator."
        ),
    },

    "Internal Audit": {
        "core_skills": [
            "Internal Audit", "Process Audit", "Stock Audit",
            "Compliance Audit", "Risk Assessment", "Financial Controls",
            "Audit Planning", "Audit Reporting",
        ],
        "bonus_skills": [
            "CIA (Certified Internal Auditor)", "ACCA", "SAP Audit Trails",
            "Data Analytics for Audit", "Forensic Accounting",
            "ERP Audit", "Bangladesh Tax Law",
        ],
        "anti_skills": ["Machine Operation", "Field Sales", "Network Administration"],
        "scoring_note": (
            "For Internal Audit: CIA or ACCA qualifications are strong education "
            "differentiators. Experience auditing FMCG or manufacturing operations "
            "is preferred over pure financial audit."
        ),
    },

    "Legal Affairs": {
        "core_skills": [
            "Contract Drafting & Review", "Legal Documentation",
            "Corporate Governance", "RJSC Compliance", "Regulatory Compliance",
            "Intellectual Property", "Due Diligence", "Labor Law",
        ],
        "bonus_skills": [
            "BIDA Compliance", "Company Secretarial Work",
            "Industrial Relations", "Factory Act Compliance",
            "Competition Law", "Litigation Management",
        ],
        "anti_skills": ["Machine Operation", "Video Editing", "Supply Chain"],
        "scoring_note": (
            "LLB from a reputable university plus Bar Council enrollment is essential. "
            "Experience advising FMCG or manufacturing companies is preferred. "
            "Knowledge of Bangladesh Companies Act 1994 and Labour Act 2006 is key."
        ),
    },

    "Admin": {
        "core_skills": [
            "Office Management", "Facility Management", "Vendor Management",
            "Document Control", "Scheduling & Calendar Management",
            "Travel & Logistics Coordination", "MS Office",
        ],
        "bonus_skills": [
            "Contract Management", "Event Management", "Record Keeping",
            "ERP Support", "Advanced Excel", "Reporting",
        ],
        "anti_skills": ["PLC Programming", "Machine Learning", "Audit"],
        "scoring_note": (
            "Admin roles value organisational reliability and multi-tasking. "
            "Candidates who have managed large offices (100+ staff) or "
            "multi-location operations score higher on experience_score."
        ),
    },

    "Management Information System (MIS)": {
        "core_skills": [
            "MIS Reporting", "Advanced Excel", "KPI Reporting",
            "Power BI", "Data Analysis", "Dashboard Development",
            "Report Automation", "Database Management",
        ],
        "bonus_skills": [
            "SQL", "Python", "Tableau", "SAP Reporting",
            "ETL Processes", "VBA / Macros",
        ],
        "anti_skills": ["Video Editing", "Machine Operation", "HR Policy"],
        "scoring_note": (
            "MIS roles require strong Excel and BI tool proficiency. "
            "Experience building real-time KPI dashboards for senior management "
            "is a strong differentiator."
        ),
    },

    # ── SCORE-04: New department profiles ──────────────────────────────────

    "Software Designer": {
        "core_skills": [
            "UI/UX Design", "Figma", "Adobe XD", "Wireframing", "Prototyping",
            "User Research", "Design Systems", "Responsive Design", "HTML/CSS",
            "Interaction Design",
        ],
        "bonus_skills": [
            "React", "Flutter", "Adobe Illustrator", "Motion Design",
            "Accessibility (WCAG)", "A/B Testing", "Usability Testing",
            "Product Thinking", "Design Tokens", "Framer",
        ],
        "anti_skills": ["PLC Programming", "Boiler Operation", "Tax Compliance", "Manual Bookkeeping"],
        "scoring_note": (
            "For Software/UI-UX Designer roles: a strong portfolio of shipped digital products "
            "is the primary evidence. Figma or Adobe XD proficiency is mandatory — candidates "
            "without a design tool score a maximum of 50 on skills_score. "
            "Front-end coding ability (HTML/CSS/React) is a strong differentiator for a "
            "digital-first company like Olympic. User research and usability testing evidence "
            "separates junior from senior designers."
        ),
    },

    "Marketing Designer": {
        "core_skills": [
            "Adobe Premiere Pro", "Adobe After Effects", "Adobe Photoshop",
            "Canva", "Video Editing", "Motion Graphics", "Graphic Design",
            "Social Media Content", "Brand Identity", "Storyboarding",
            "Color Theory", "Typography",
        ],
        "bonus_skills": [
            "Adobe Illustrator", "Adobe InDesign", "3D Animation (Blender/Cinema4D)",
            "YouTube/TikTok content creation", "Meta Ads creative",
            "Packaging Design", "New Product Launch visuals",
            "Google Analytics (content performance)", "Drone footage editing",
        ],
        "anti_skills": ["PLC Programming", "Tax Compliance", "Supply Chain", "Audit"],
        "scoring_note": (
            "For Marketing Designer / Content Creator roles: Adobe Premiere Pro AND After Effects "
            "are both mandatory for video-heavy roles. A candidate missing BOTH caps at 50 on skills_score. "
            "Portfolio quality matters more than years of experience — evidence of consumer brand "
            "campaigns (FMCG / food / beverage) is a strong differentiator. "
            "Social media metric evidence (views, engagement rates, follower growth) should increase "
            "experience_score. For pure graphic design roles (no video), Photoshop + Illustrator "
            "are the mandatory tools."
        ),
    },

    "Institutional Sales": {
        "core_skills": [
            "Institutional Sales", "B2B Sales", "Tender Management",
            "Key Account Management", "Contract Negotiation",
            "Government/NGO/Corporate Client Management",
            "Sales Reporting", "Revenue Target Achievement",
        ],
        "bonus_skills": [
            "Trade Finance", "Letter of Credit (LC)", "Credit Management",
            "CRM Tools", "SAP SD", "Sales Forecasting", "Pipeline Management",
            "New Business Development", "FMCG Institutional Channel",
        ],
        "anti_skills": ["PLC Programming", "Video Editing", "Machine Learning"],
        "scoring_note": (
            "Institutional Sales is distinct from field FMCG retail sales. "
            "Candidates should show experience with large-ticket B2B or institutional accounts "
            "(hospitals, hotels, government canteens, NGOs). "
            "Quantified revenue achievements (BDT or USD targets met/exceeded) are mandatory "
            "for Shortlist consideration. "
            "Field FMCG retail experience alone does not qualify for institutional sales roles."
        ),
    },

    "Field Force": {
        "core_skills": [
            # Field Force Operations (admin/coordination role from BDJobs)
            "Advanced Microsoft Excel", "Field Force Database Management",
            "Attendance & Leave Records", "Recruitment Coordination",
            "Reporting & Dashboard Preparation",
            # Traditional field skills (for field sales sub-roles)
            "Field Sales", "Beat Planning", "Outlet Coverage",
            "Distributor Management", "Secondary Sales",
        ],
        "bonus_skills": [
            "ERP/SAP/SFA Systems", "MBA", "Field Force Management Experience",
            "Sales Administration", "Van Sales", "New Outlet Opening",
        ],
        "anti_skills": ["Software Engineering", "Financial Audit", "Machine Learning"],
        "scoring_note": (
            "The Executive - Field Force Operations role at Olympic is an OPERATIONS/COORDINATION role "
            "supporting the field team \u2014 NOT a direct sales role. "
            "Advanced Excel proficiency is the PRIMARY skill for this role. "
            "HR documentation, attendance tracking, and reporting are core functions. "
            "Freshers with strong Excel and BBA background are explicitly welcome for this role."
        ),
    },

    "VAT / VAT & Delivery": {
        "core_skills": [
            "VAT Compliance", "Bangladesh VAT Law (VAT & SD Act 2012)",
            "Mushak Forms (Mushak 6.1, 6.2, 6.3, etc.)", "VAT Return Filing",
            "VAT Audit", "Tax Compliance", "Financial Reporting",
            "Tally / Accounting Software", "MS Excel",
        ],
        "bonus_skills": [
            "SAP FICO", "Income Tax", "NBR Correspondence",
            "VAT Appeal Procedures", "Transfer Pricing",
            "Internal Audit", "ERP VAT Module",
        ],
        "anti_skills": ["Machine Learning", "Video Editing", "PLC Programming"],
        "scoring_note": (
            "VAT roles in Bangladesh require deep knowledge of the VAT & Supplementary Duty Act 2012 "
            "and proficiency with Mushak forms. Candidates without this specific knowledge "
            "should score no higher than 50 on skills_score. NBR audit experience is a strong signal. "
            "For combined VAT & Delivery roles, distribution/logistics exposure is also required."
        ),
    },

    "Operations": {
        "core_skills": [
            "Operations Management", "Process Improvement", "KPI Management",
            "Cross-functional Coordination", "SOP Development",
            "Resource Planning", "Budget Management", "Reporting & Analytics",
            "Vendor Management", "Compliance",
        ],
        "bonus_skills": [
            "Lean / Six Sigma", "Project Management (PMP)", "SAP",
            "Business Process Reengineering", "Change Management",
            "FMCG Operations", "Multi-site Operations",
        ],
        "anti_skills": ["Video Editing", "PLC Programming", "Tax Filing"],
        "scoring_note": (
            "Operations roles at Olympic span factory, distribution, and corporate functions. "
            "Candidates should show cross-functional coordination at scale. "
            "FMCG or manufacturing operations background is strongly preferred. "
            "Evidence of measurable process improvements (cost savings, cycle time reduction) "
            "strongly increases experience_score."
        ),
    },

    "Corporate Affairs": {
        "core_skills": [
            "Corporate Governance", "RJSC Compliance", "Company Secretarial Work",
            "Board Meeting Management", "Regulatory Compliance",
            "Shareholder Relations", "Annual Report Preparation",
            "Legal Documentation", "Bangladesh Companies Act 1994",
        ],
        "bonus_skills": [
            "BSEC Regulations", "DSEX Listed Company Compliance",
            "AGM/EGM Coordination", "Due Diligence",
            "MoU / JV Agreement Management", "BIDA Compliance",
            "ACS / FCS (Fellow Company Secretary) qualification",
        ],
        "anti_skills": ["Machine Operation", "Video Editing", "Field Sales"],
        "scoring_note": (
            "For Corporate Affairs / Company Secretary roles: "
            "knowledge of Bangladesh Companies Act 1994, Securities Act, and BSEC guidelines is mandatory. "
            "ACS (Associate Company Secretary) or FCS qualification from ICSB (Institute of Chartered "
            "Secretaries of Bangladesh) is a strong education differentiator. "
            "Experience with a DSE-listed company is a strong positive signal."
        ),
    },

    "Security": {
        "core_skills": [
            "Access Control Systems", "Fire Safety & Alarm Systems",
            "Incident Reporting & Documentation", "Security Equipment Handling",
            "Risk Assessment & Mitigation", "CCTV Surveillance",
            "Guard Management", "Factory Security Operations",
            "Patrol Management", "Visitor & Vehicle Control",
        ],
        "bonus_skills": [
            "Ex-Defense / Military Background",
            "ISO/HACCP/BSCI/SEDEX/WRAP Compliance Audit Support",
            "Fire Safety Certification",
            "Emergency Response Training",
            "Loss Prevention",
        ],
        "anti_skills": [
            "Software Development", "Financial Reporting", "Marketing",
            "Machine Learning", "Tax Compliance",
        ],
        "scoring_note": (
            "For Security roles at Olympic Industries: factory/industrial security experience is "
            "mandatory \u2014 mall, bank, or event security alone is insufficient. "
            "Ex-Defense background is explicitly preferred and should increase culture_fit_score by ~10. "
            "Compliance audit support (ISO, HACCP, BSCI, SEDEX, WRAP) is a key differentiator "
            "for the Security Officer role and should increase skills_score. "
            "Basic computer literacy (MS Office, CCTV software) is required. "
            "Leadership evidence (managing guards/shift team) is scored under leadership_score "
            "and should reflect Level 2-4 depending on seniority."
        ),
    },

    "Market Audit": {
        "core_skills": [
            "Market Visits & Gap Analysis", "Stock Counting & Reconciliation",
            "Product Presence Monitoring", "Market Share Analysis",
            "Audit Report Preparation", "FMCG Channel Knowledge",
            "ERP", "MS Office",
        ],
        "bonus_skills": [
            "Distributor/Dealer Audit", "Competitor Analysis",
            "SKU-wise Availability Survey", "Marketing Campaign Audit",
        ],
        "anti_skills": ["Machine Learning", "Software Engineering", "Electrical Engineering"],
        "scoring_note": (
            "Market Audit is a field-intensive role requiring frequent travel to trade channels. "
            "Willingness to travel is mandatory \u2014 candidates who indicate unwillingness should be flagged. "
            "FMCG channel knowledge (distributor, retail, dealer) is required. "
            "Marketing campaign audit is an unusual combined function \u2014 both field and analytical skills needed."
        ),
    },
}

# Departments without a specific profile fall back to generic scoring.

RED_FLAG_PRESETS = [
    # Stability
    "Frequent job changes",
    "Short tenures (<1yr)",
    "Employment gaps",
    "Career regression",
    "Over-reliance on one employer",

    # Qualification
    "No relevant degree",
    "Missing certifications",
    "Overqualified",
    "Underqualified",
    "Skills mismatch",

    # Experience
    "No FMCG experience",
    "No industry experience",
    "No leadership experience",
    "No measurable achievements",
    "Vague job descriptions",
    "Limited local market knowledge",

    # Red signals
    "No SAP/ERP exposure",
    "No references listed",
    "Salary expectations too high",
    "Relocation required",
    "Poor communication indicators",
]

SCORE_DIMS = [
    ("overall_score",     "Overall"),
    ("skills_score",      "Skills"),
    ("experience_score",  "Experience"),
    ("leadership_score",  "Leadership"),
    ("education_score",   "Education"),
    ("culture_fit_score", "Culture Fit"),
]

SCORE_COLS = SCORE_DIMS  # legacy alias

VERDICT_CFG = {
    "Shortlist": {"color":"#16A34A","bg":"#DCFCE7","icon":"🟢","dark_bg":"#14532D","dark_color":"#86EFAC"},
    "Maybe":     {"color":"#D97706","bg":"#FEF3C7","icon":"🟡","dark_bg":"#451A03","dark_color":"#FCD34D"},
    "Reject":    {"color":"#DC2626","bg":"#FEE2E2","icon":"🔴","dark_bg":"#450A0A","dark_color":"#FCA5A5"},
}

VERDICT_STYLE = {
    k: {"icon": v["icon"], "color_dark": v["dark_color"], "color_light": v["color"]}
    for k, v in VERDICT_CFG.items()
}

RED      = "#C8102E"
RED_DARK = "#A00D24"
RED_PALE = "#FEE2E2"

# ── Schema ─────────────────────────────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS jobs (
    id                SERIAL PRIMARY KEY,
    job_label         TEXT UNIQUE NOT NULL,
    job_title         TEXT,
    department        TEXT,
    jd_text           TEXT,
    required_skills   TEXT[],
    red_flags         TEXT[],
    min_experience    TEXT,
    education_req     TEXT,
    weight_skills     INTEGER DEFAULT 50,
    weight_exp        INTEGER DEFAULT 30,
    weight_edu        INTEGER DEFAULT 20,
    weight_leadership INTEGER DEFAULT 10,
    weight_culture    INTEGER DEFAULT 5,
    interviewer_notes TEXT,
    status            TEXT DEFAULT 'Pending',
    created_at        TIMESTAMP DEFAULT NOW(),
    last_ranked_at    TIMESTAMP
);

CREATE TABLE IF NOT EXISTS candidates (
    id                SERIAL PRIMARY KEY,
    job_label         TEXT NOT NULL,
    apply_id          TEXT NOT NULL,
    candidate_name    TEXT,
    email             TEXT,
    mobile            TEXT,
    location          TEXT,
    degree            TEXT,
    university        TEXT,
    experience_detail TEXT,
    age               NUMERIC(5,1),
    expected_salary   TEXT,
    current_salary    TEXT,
    application_date  TEXT,
    bdjobs_score      TEXT,
    has_uploaded_cv   BOOLEAN DEFAULT FALSE,
    profile_txt_path  TEXT,
    pdf_path          TEXT,
    pdf_text_chars    INTEGER DEFAULT 0,
    jd_used           TEXT DEFAULT '',
    overall_score     INTEGER,
    skills_score      INTEGER,
    experience_score  INTEGER,
    leadership_score  INTEGER,
    education_score   INTEGER,
    edu_tier_score    INTEGER,
    edu_degree_score  INTEGER,
    edu_gpa_score     INTEGER,
    culture_fit_score INTEGER,
    experience_years  NUMERIC(4,1),
    strengths         TEXT[],
    gaps              TEXT[],
    risk_flags        TEXT[],
    recommendation    TEXT CHECK (recommendation IN ('Shortlist','Maybe','Reject')),
    reasoning         TEXT,
    hr_override       TEXT,
    hr_note           TEXT,
    ranked_at         TIMESTAMP DEFAULT NOW(),
    rank_error        TEXT,
    UNIQUE (job_label, apply_id)
);

-- HR override audit log (Part 7.5) — every time HR changes a recommendation,
-- a row is appended here.  Required for compliance / fairness audit trails.
CREATE TABLE IF NOT EXISTS hr_audit_log (
    id          SERIAL PRIMARY KEY,
    job_label   TEXT NOT NULL,
    apply_id    TEXT NOT NULL,
    hr_user     TEXT,
    old_value   TEXT,
    new_value   TEXT,
    note        TEXT,
    changed_at  TIMESTAMP DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_c_job   ON candidates(job_label);
CREATE INDEX IF NOT EXISTS idx_c_rec   ON candidates(recommendation);
CREATE INDEX IF NOT EXISTS idx_c_score ON candidates(overall_score DESC);
CREATE INDEX IF NOT EXISTS idx_c_email ON candidates(email);
CREATE INDEX IF NOT EXISTS idx_c_mobile ON candidates(mobile);
CREATE INDEX IF NOT EXISTS idx_j_label ON jobs(job_label);
CREATE INDEX IF NOT EXISTS idx_j_dept  ON jobs(department);
CREATE INDEX IF NOT EXISTS idx_hr_audit_job ON hr_audit_log(job_label);
CREATE INDEX IF NOT EXISTS idx_hr_audit_app ON hr_audit_log(apply_id);

-- BDJobs credentials table — stores recruiter login for auto-login
CREATE TABLE IF NOT EXISTS bdjobs_credentials (
    id            SERIAL PRIMARY KEY,
    username      TEXT NOT NULL,
    password      TEXT NOT NULL,  -- base64-encoded obfuscation (not true encryption)
    created_at    TIMESTAMP DEFAULT NOW(),
    updated_at    TIMESTAMP DEFAULT NOW()
);
"""

# Additive migrations that run every startup — safe to re-run.
MIGRATION_SQL = """
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS updated_at        TIMESTAMP DEFAULT NOW();
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS weight_leadership INTEGER DEFAULT 10;
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS weight_culture    INTEGER DEFAULT 5;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS edu_tier_score   INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS edu_degree_score INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS edu_gpa_score    INTEGER;

-- PHASE 5: raw-score snapshots written by ranker.normalise_job_scores so the
-- pre-normalisation values stay auditable. Created lazily by the ranker when
-- it first runs --normalise; declaring them here lets fresh installs migrate
-- without a separate pass.
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS overall_raw      INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS skills_raw       INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS experience_raw   INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS leadership_raw   INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS education_raw    INTEGER;
ALTER TABLE candidates ADD COLUMN IF NOT EXISTS culture_fit_raw  INTEGER;

-- BDJobs salary & metadata columns
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS salary_stated    TEXT;
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS salary_estimate  TEXT;
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS location         TEXT;
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS deadline         TEXT;
ALTER TABLE jobs ADD COLUMN IF NOT EXISTS bdjobs_experience TEXT;
ALTER TABLE candidates ALTER COLUMN rank_error TYPE TEXT;   -- remove any length limit

-- PHASE 5: indexes used by FEAT-01 (compare candidates) and the dashboard.
CREATE INDEX IF NOT EXISTS idx_c_apply_id ON candidates(apply_id);
CREATE INDEX IF NOT EXISTS idx_c_ranked   ON candidates(ranked_at DESC);

CREATE TABLE IF NOT EXISTS bdjobs_credentials (
    id            SERIAL PRIMARY KEY,
    username      TEXT NOT NULL,
    password      TEXT NOT NULL,
    created_at    TIMESTAMP DEFAULT NOW(),
    updated_at    TIMESTAMP DEFAULT NOW()
);

UPDATE jobs SET department = 'Uncategorized'
  WHERE department IS NULL OR department = '';

-- Authentication & audit tables
CREATE TABLE IF NOT EXISTS users (
    id            SERIAL PRIMARY KEY,
    username      TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    display_name  TEXT,
    role          TEXT NOT NULL DEFAULT 'user' CHECK (role IN ('admin','user')),
    is_active     BOOLEAN DEFAULT TRUE,
    created_at    TIMESTAMP DEFAULT NOW(),
    last_login    TIMESTAMP
);

CREATE TABLE IF NOT EXISTS audit_logs (
    id          SERIAL PRIMARY KEY,
    user_id     INTEGER REFERENCES users(id) ON DELETE SET NULL,
    username    TEXT,
    action      TEXT NOT NULL,
    target_type TEXT,
    target_id   TEXT,
    details     TEXT,
    ip_address  TEXT,
    created_at  TIMESTAMP DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_logs(user_id);
CREATE INDEX IF NOT EXISTS idx_audit_action ON audit_logs(action);
CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_logs(created_at DESC);

-- User profile fields
ALTER TABLE users ADD COLUMN IF NOT EXISTS department  TEXT;
ALTER TABLE users ADD COLUMN IF NOT EXISTS email       TEXT;
ALTER TABLE users ADD COLUMN IF NOT EXISTS phone       TEXT;
ALTER TABLE users ADD COLUMN IF NOT EXISTS employee_id TEXT;
"""

# ── Connection ─────────────────────────────────────────────────────────────────

def _new_conn():
    """Create a fresh connection and ensure schema + migrations + backfill."""
    if not pg_is_configured():
        raise RuntimeError(
            "PostgreSQL is not configured.\n\n"
            "For Streamlit Cloud: add your Neon DB credentials in the app secrets\n"
            "(Settings → Secrets). Required fields: PG_HOST, PG_USER, PG_PASSWORD.\n\n"
            "For local use: set PG_HOST=localhost, PG_USER=postgres, PG_PASSWORD=yourpassword "
            "in a .env file or environment variables."
        )
    conn = psycopg2.connect(**PG_CONN)
    conn.autocommit = True
    ensure_schema(conn)
    return conn


def ensure_schema(conn) -> None:
    """Create tables, run migrations, and backfill `jobs` rows from any
    candidate `job_label` not yet registered.  Idempotent and cheap.
    """
    with conn.cursor() as cur:
        cur.execute("SET search_path = public")
        cur.execute(SCHEMA_SQL)
        cur.execute(MIGRATION_SQL)
        # Backfill: any candidate.job_label with no jobs row → Uncategorized.
        cur.execute("""
            INSERT INTO jobs (job_label, department)
            SELECT DISTINCT c.job_label, 'Uncategorized'
              FROM candidates c
              LEFT JOIN jobs j ON j.job_label = c.job_label
             WHERE j.job_label IS NULL
            ON CONFLICT (job_label) DO NOTHING
        """)
    # NEW: seed BDJobs registry (idempotent)
    seed_bdjobs_registry(conn)


def seed_bdjobs_registry(conn=None) -> int:
    """Insert all live BDJobs listings into the jobs table.

    Uses ON CONFLICT DO NOTHING so existing ranked jobs are not overwritten.
    Returns count of newly inserted rows. Safe to call multiple times — idempotent.
    """
    if conn is None:
        conn = fresh_conn()

    inserted = 0
    for job_label, meta in BDJOBS_JOB_REGISTRY.items():
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO jobs (
                        job_label, job_title, department,
                        required_skills, red_flags,
                        min_experience, education_req,
                        weight_skills, weight_exp, weight_edu,
                        weight_leadership, weight_culture,
                        interviewer_notes, status,
                        salary_stated, salary_estimate, location, deadline, bdjobs_experience
                    ) VALUES (
                        %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        %s, 'Pending',
                        %s, %s, %s, %s, %s
                    )
                    ON CONFLICT (job_label) DO NOTHING
                """, (
                    job_label,
                    meta.get("job_title", ""),
                    meta.get("department", "Uncategorized"),
                    meta.get("required_skills", []),
                    meta.get("red_flags", []),
                    meta.get("min_experience", "Any"),
                    meta.get("education_req", "Any"),
                    meta.get("weight_skills", 50),
                    meta.get("weight_exp", 30),
                    meta.get("weight_edu", 10),
                    meta.get("weight_leadership", 5),
                    meta.get("weight_culture", 5),
                    meta.get("scoring_note", ""),
                    meta.get("salary_stated", "Negotiable"),
                    meta.get("salary_estimate", ""),
                    meta.get("location", ""),
                    meta.get("deadline", ""),
                    meta.get("experience", ""),
                ))
            inserted += 1
        except Exception as e:
            print(f"[seed_bdjobs_registry] Failed for {job_label}: {e}")
    return inserted


def migrate_existing_jobs(conn) -> int:
    """Explicit wrapper for the backfill — returns rows inserted."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO jobs (job_label, department)
            SELECT DISTINCT c.job_label, 'Uncategorized'
              FROM candidates c
              LEFT JOIN jobs j ON j.job_label = c.job_label
             WHERE j.job_label IS NULL
            ON CONFLICT (job_label) DO NOTHING
            RETURNING job_label
        """)
        return len(cur.fetchall())


def get_conn():
    """Return a live psycopg2 connection, reconnecting if stale/closed.

    psycopg2 connections held in @st.cache_resource go stale when Postgres
    restarts or when the session times out. We keep one per-session and
    transparently reconnect on failure.
    """
    try:
        conn = st.session_state.get("pg_conn")
        if conn is None or getattr(conn, "closed", 1):
            conn = _new_conn()
            st.session_state["pg_conn"] = conn
        else:
            # Liveness ping — dead connections raise here.
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
        return conn
    except Exception:
        conn = _new_conn()
        st.session_state["pg_conn"] = conn
        return conn


def fresh_conn():
    """Always returns a brand-new autocommit connection. Caller closes it."""
    return _new_conn()

# ── Cross-page processing status (filesystem-based, no session_state) ─────────

def get_active_processing() -> list[dict]:
    """Scan downloaded_resumes/*/_ranker_progress.jsonl for live ranker runs.

    Returns list of dicts with keys: job, total, processed, errors, last_ts,
    is_running (bool). A run is considered running if there is no `done` event
    AND the last event timestamp is within the last 60 seconds.
    """
    from datetime import datetime as _dt, timedelta as _td
    from pathlib import Path as _P

    base = _P(os.environ.get("RESUMES_BASE", str(_P(__file__).resolve().parent.parent / "downloaded_resumes")))
    out: list[dict] = []
    if not base.exists():
        return out

    now = _dt.now()
    for jdir in base.iterdir():
        if not jdir.is_dir():
            continue
        prog = jdir / "_ranker_progress.jsonl"
        if not prog.exists():
            continue
        try:
            total = processed = errors = 0
            last_ts = None
            done = False
            with open(prog, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        ev = json.loads(line)
                    except Exception:
                        continue
                    e = ev.get("event")
                    if e == "start":
                        total = ev.get("total", 0)
                    elif e in ("ok", "ok_fallback", "error"):
                        processed += 1
                        if e == "error":
                            errors += 1
                        if "ts" in ev:
                            last_ts = ev["ts"]
                    elif e == "done":
                        done = True
                        if "ts" in ev:
                            last_ts = ev["ts"]
            is_running = False
            if not done and last_ts:
                try:
                    delta = now - _dt.fromisoformat(last_ts)
                    is_running = delta < _td(seconds=120)
                except Exception:
                    pass
            out.append({
                "job": jdir.name,
                "total": total,
                "processed": processed,
                "errors": errors,
                "last_ts": last_ts,
                "is_running": is_running,
                "done": done,
            })
        except Exception:
            continue
    # Active first
    out.sort(key=lambda x: (not x["is_running"], x["job"]))
    return out


def render_processing_banner() -> None:
    """Render a small banner at the top of any page indicating active processing.

    Safe to call on any page; renders nothing when no jobs are being processed.
    """
    active = [r for r in get_active_processing() if r["is_running"]]
    if not active:
        return
    msgs = []
    for r in active:
        pct = (r["processed"] / r["total"] * 100.0) if r["total"] else 0
        msgs.append(
            f"**{r['job']}** — {r['processed']}/{r['total'] or '?'} ranked ({pct:.0f}%)"
            + (f" · ⚠ {r['errors']} errors" if r["errors"] else "")
        )
    st.info(
        "⚡ **Live ranking in progress** (data refreshes automatically below):  \n"
        + "  \n".join(msgs)
        + "\n\nYou can navigate freely — rankings update live as candidates are processed.",
        icon="🔄",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Authentication helpers
# ═══════════════════════════════════════════════════════════════════════════════

try:
    import bcrypt
except Exception:  # pragma: no cover
    bcrypt = None


def _hash_password(password: str) -> str:
    """Hash a plain-text password using bcrypt."""
    if bcrypt is None:
        raise RuntimeError("bcrypt is not installed. Run: pip install bcrypt")
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")


def _verify_password(password: str, password_hash: str) -> bool:
    """Verify a plain-text password against a bcrypt hash."""
    if bcrypt is None:
        return False
    try:
        return bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8"))
    except Exception:
        return False


def authenticate_user(conn, username: str, password: str) -> dict | None:
    """Validate credentials. Returns user dict {id, username, display_name, role}
    on success, None on failure."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, username, password_hash, display_name, role, is_active FROM users WHERE username = %s",
            (username,),
        )
        row = cur.fetchone()
    if not row:
        return None
    uid, uname, pwd_hash, dname, role, is_active = row
    if not is_active:
        return None
    if not _verify_password(password, pwd_hash):
        return None
    # Update last_login
    with conn.cursor() as cur:
        cur.execute("UPDATE users SET last_login = NOW() WHERE id = %s", (uid,))
    return {"id": uid, "username": uname, "display_name": dname or uname, "role": role}


def create_user(conn, username: str, password: str, display_name: str | None = None,
                role: str = "user", created_by: str | None = None,
                department: str | None = None, email: str | None = None,
                phone: str | None = None, employee_id: str | None = None) -> tuple[bool, str]:
    """Create a new user. Returns (success, message)."""
    if not username or not password:
        return False, "Username and password are required."
    if len(password) < 6:
        return False, "Password must be at least 6 characters."
    if role not in ("admin", "user"):
        return False, "Invalid role."
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO users
                    (username, password_hash, display_name, role, department, email, phone, employee_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (username, _hash_password(password), display_name or username, role,
                 department, email, phone, employee_id),
            )
        return True, f"User '{username}' created successfully."
    except psycopg2.IntegrityError:
        return False, f"Username '{username}' already exists."
    except Exception as e:
        return False, f"Error creating user: {e}"


def list_users(conn) -> list[dict]:
    """Return all users ordered by creation date."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, username, display_name, role, is_active, department, email, phone, employee_id, "
            "created_at, last_login FROM users ORDER BY created_at DESC"
        )
        cols = [desc[0] for desc in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def toggle_user_active(conn, user_id: int, is_active: bool) -> None:
    """Enable or disable a user account."""
    with conn.cursor() as cur:
        cur.execute("UPDATE users SET is_active = %s WHERE id = %s", (is_active, user_id))


def log_audit(conn, user_id: int | None, username: str | None, action: str,
              target_type: str | None = None, target_id: str | None = None,
              details: str | None = None) -> None:
    """Write an entry to the audit log."""
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO audit_logs (user_id, username, action, target_type, target_id, details) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                (user_id, username, action, target_type, target_id, details),
            )
    except Exception:
        pass  # Never block the UI for audit logging


def get_audit_logs(conn, limit: int = 500) -> list[dict]:
    """Return recent audit logs."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, username, action, target_type, target_id, details, created_at "
            "FROM audit_logs ORDER BY created_at DESC LIMIT %s",
            (limit,),
        )
        cols = [desc[0] for desc in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


# ── Metadata ingestion ─────────────────────────────────────────────────────────

def ingest_metadata(job_label: str, meta_csv: str) -> tuple:
    if not os.path.exists(meta_csv):
        return 0, 0
    conn = fresh_conn()
    updated = skipped = 0
    with open(meta_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            apply_id = str(row.get("apply_id") or row.get("ApplyID") or "").strip()
            if not apply_id:
                skipped += 1
                continue
            pdf_file = str(row.get("uploaded_cv_file") or "").strip()
            pdf_path = os.path.join(RESUMES_BASE, job_label, "uploaded_cvs", pdf_file) if pdf_file else ""
            try:
                age = float(str(row.get("age") or "0").replace(",",""))
            except ValueError:
                age = None
            has_cv = str(row.get("has_uploaded_cv") or "").strip().lower() in ("yes","true","1")
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO candidates
                        (job_label, apply_id, candidate_name, email, mobile, location,
                         degree, university, experience_detail, age,
                         expected_salary, current_salary, application_date,
                         bdjobs_score, has_uploaded_cv, pdf_path)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (job_label, apply_id) DO UPDATE SET
                        candidate_name    = EXCLUDED.candidate_name,
                        email             = EXCLUDED.email,
                        mobile            = EXCLUDED.mobile,
                        location          = EXCLUDED.location,
                        degree            = EXCLUDED.degree,
                        university        = EXCLUDED.university,
                        experience_detail = EXCLUDED.experience_detail,
                        age               = EXCLUDED.age,
                        expected_salary   = EXCLUDED.expected_salary,
                        current_salary    = EXCLUDED.current_salary,
                        application_date  = EXCLUDED.application_date,
                        bdjobs_score      = EXCLUDED.bdjobs_score,
                        has_uploaded_cv   = EXCLUDED.has_uploaded_cv,
                        pdf_path          = COALESCE(EXCLUDED.pdf_path, candidates.pdf_path)
                """, (
                    job_label, apply_id,
                    str(row.get("candidate_name") or "").strip(),
                    str(row.get("email") or "").strip(),
                    str(row.get("mobile") or "").strip(),
                    str(row.get("location") or "").strip(),
                    str(row.get("degree") or "").strip(),
                    str(row.get("university") or "").strip(),
                    str(row.get("experience") or "").strip(),
                    age,
                    str(row.get("expected_salary") or "").strip(),
                    str(row.get("current_salary") or "").strip(),
                    str(row.get("application_date") or "").strip(),
                    str(row.get("bdjobs_match_score") or "").strip(),
                    has_cv, pdf_path,
                ))
            updated += 1
    conn.close()
    return updated, skipped

# ── Job queries ────────────────────────────────────────────────────────────────

def fetch_all_jobs(conn) -> pd.DataFrame:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                j.job_label, j.job_title, j.department, j.status,
                j.created_at, j.last_ranked_at,
                COUNT(c.id)                                                   AS total,
                SUM(CASE WHEN c.overall_score IS NOT NULL THEN 1 ELSE 0 END) AS ranked,
                SUM(CASE WHEN c.recommendation='Shortlist' THEN 1 ELSE 0 END) AS shortlisted,
                ROUND(AVG(c.overall_score))                                   AS avg_score
            FROM jobs j
            LEFT JOIN candidates c ON c.job_label = j.job_label
            GROUP BY j.job_label, j.job_title, j.department, j.status,
                     j.created_at, j.last_ranked_at
            HAVING COUNT(c.id) > 0
            ORDER BY j.created_at DESC
        """)
        rows = cur.fetchall()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["created_at"]     = pd.to_datetime(df["created_at"]).dt.strftime("%d %b %Y")
    df["last_ranked_at"] = pd.to_datetime(df["last_ranked_at"]).dt.strftime("%d %b %Y %H:%M")
    return df


def list_download_folders() -> list:
    """Return one entry per folder under RESUMES_BASE with download stats.

    Each entry is a dict:
        {
            "name": str,                    # folder name (== job_label)
            "path": str,                    # absolute folder path
            "n_profiles": int,              # # .txt files in profiles_txt/
            "n_cvs": int,                   # # PDFs in uploaded_cvs/
            "size_mb": float,               # total recursive size, MB
            "modified": datetime|None,      # mtime of most recent file
            "has_metadata_csv": bool,       # whether <name>_metadata.csv exists
        }
    Sorted newest-first by modified time.
    """
    base = Path(RESUMES_BASE)
    out = []
    if not base.is_dir():
        return out
    for folder in sorted([p for p in base.iterdir() if p.is_dir()]):
        profiles_dir = folder / "profiles_txt"
        cvs_dir      = folder / "uploaded_cvs"
        n_profiles = len([f for f in profiles_dir.glob("*.txt")]) if profiles_dir.is_dir() else 0
        n_cvs      = len([f for f in cvs_dir.glob("*.pdf")])      if cvs_dir.is_dir() else 0
        total_bytes = 0
        latest_mtime = 0.0
        for sub in folder.rglob("*"):
            if sub.is_file():
                try:
                    s = sub.stat()
                    total_bytes += s.st_size
                    if s.st_mtime > latest_mtime:
                        latest_mtime = s.st_mtime
                except OSError:
                    pass
        out.append({
            "name":              folder.name,
            "path":              str(folder),
            "n_profiles":        n_profiles,
            "n_cvs":             n_cvs,
            "size_mb":           round(total_bytes / (1024 * 1024), 2),
            "modified":          datetime.fromtimestamp(latest_mtime) if latest_mtime else None,
            "has_metadata_csv":  (folder / f"{folder.name}_metadata.csv").is_file(),
        })
    out.sort(key=lambda r: r["modified"] or datetime.min, reverse=True)
    return out


def fetch_departments_with_roles(conn) -> list[dict]:
    """
    Returns all departments that have at least one job posting,
    grouped with their roles and live applicant/ranked/error counts.

    Called by the Job Rankings landing accordion.
    """
    sql = """
        SELECT
            COALESCE(j.department, 'Uncategorized')          AS department,
            j.job_label,
            j.job_title,
            j.status,
            j.location,
            j.min_experience,
            j.education_req,
            j.required_skills,
            COUNT(c.id)                                       AS total,
            SUM(CASE WHEN c.overall_score IS NOT NULL
                     THEN 1 ELSE 0 END)                       AS ranked,
            SUM(CASE WHEN c.recommendation = 'Shortlist'
                     THEN 1 ELSE 0 END)                       AS shortlisted,
            SUM(CASE WHEN c.recommendation = 'Maybe'
                     THEN 1 ELSE 0 END)                       AS maybe,
            SUM(CASE WHEN c.recommendation = 'Reject'
                     THEN 1 ELSE 0 END)                       AS rejected,
            SUM(CASE WHEN c.rank_error IS NOT NULL
                          AND c.overall_score IS NULL
                     THEN 1 ELSE 0 END)                       AS errors,
            ROUND(AVG(c.overall_score))                       AS avg_score
        FROM jobs j
        LEFT JOIN candidates c ON c.job_label = j.job_label
        GROUP BY
            j.department, j.job_label, j.job_title, j.status,
            j.location, j.min_experience,
            j.education_req, j.required_skills, j.created_at
        HAVING COUNT(c.id) > 0
        ORDER BY j.department ASC, j.created_at DESC NULLS LAST
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    from collections import defaultdict
    dept_map = defaultdict(lambda: {
        "department":        "",
        "roles":             [],
        "total_roles":       0,
        "total_applicants":  0,
        "total_ranked":      0,
        "total_shortlisted": 0,
        "total_errors":      0,
    })

    for r in rows:
        dept = r.get("department") or "Uncategorized"
        role = {
            "job_label":       r["job_label"],
            "job_title":       r.get("job_title") or r["job_label"],
            "status":          r.get("status")          or "Pending",
            "salary_range":    r.get("salary_range")    or "Negotiable",
            "location":        r.get("location")        or "Bangladesh",
            "min_experience":  r.get("min_experience")  or "Any",
            "education_req":   r.get("education_req")   or "Any",
            "required_skills": r.get("required_skills") or [],
            "total":           int(r.get("total")       or 0),
            "ranked":          int(r.get("ranked")      or 0),
            "shortlisted":     int(r.get("shortlisted") or 0),
            "maybe":           int(r.get("maybe")       or 0),
            "rejected":        int(r.get("rejected")    or 0),
            "errors":          int(r.get("errors")      or 0),
            "avg_score":       int(r["avg_score"]) if r.get("avg_score") is not None else None,
        }
        d = dept_map[dept]
        d["department"]        = dept
        d["roles"].append(role)
        d["total_roles"]       += 1
        d["total_applicants"]  += role["total"]
        d["total_ranked"]      += role["ranked"]
        d["total_shortlisted"] += role["shortlisted"]
        d["total_errors"]      += role["errors"]

    return sorted(dept_map.values(), key=lambda x: x["department"])


def fetch_job_labels(conn) -> list:
    """Back-compat helper used by older pages.  Now also returns department.

    Tuple layout: (job_label, total, ranked, last_ranked_at, department).
    """
    df = fetch_all_jobs(conn)
    if df.empty:
        return []
    result = []
    for _, row in df.iterrows():
        result.append((
            row["job_label"],
            int(row.get("total") or 0),
            int(row.get("ranked") or 0),
            row.get("last_ranked_at"),
            str(row.get("department") or "Uncategorized"),
        ))
    return result


# ── Department queries ─────────────────────────────────────────────────────────

def fetch_departments(conn) -> list:
    """Departments that have at least one candidate (ranked or not).

    Each dict: {department, job_count, total_candidates, ranked_candidates,
    shortlist, maybe, reject, last_run}.  Sorted by ranked_candidates desc.
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                COALESCE(j.department, 'Uncategorized')               AS department,
                COUNT(DISTINCT CASE WHEN c.id IS NOT NULL
                                    THEN j.job_label END)              AS job_count,
                COUNT(c.id)                                           AS total_candidates,
                SUM(CASE WHEN c.overall_score IS NOT NULL
                         THEN 1 ELSE 0 END)                           AS ranked_candidates,
                SUM(CASE WHEN c.recommendation='Shortlist'
                         THEN 1 ELSE 0 END)                           AS shortlist,
                SUM(CASE WHEN c.recommendation='Maybe'
                         THEN 1 ELSE 0 END)                           AS maybe,
                SUM(CASE WHEN c.recommendation='Reject'
                         THEN 1 ELSE 0 END)                           AS reject,
                MAX(c.ranked_at)                                      AS last_run
            FROM jobs j
            LEFT JOIN candidates c ON c.job_label = j.job_label
            GROUP BY COALESCE(j.department, 'Uncategorized')
            HAVING COUNT(c.id) > 0
            ORDER BY ranked_candidates DESC NULLS LAST, total_candidates DESC
        """)
        rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get("last_run"):
            d["last_run"] = pd.to_datetime(d["last_run"]).strftime("%d %b %Y %H:%M")
        for k in ("job_count", "total_candidates", "ranked_candidates",
                  "shortlist", "maybe", "reject"):
            d[k] = int(d.get(k) or 0)
        result.append(d)
    return result


def fetch_jobs_by_department(conn, department: str) -> list:
    """All jobs under a department with per-verdict counts."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                j.job_label, j.job_title,
                COALESCE(j.department, 'Uncategorized')               AS department,
                COUNT(c.id)                                           AS total,
                SUM(CASE WHEN c.overall_score IS NOT NULL
                         THEN 1 ELSE 0 END)                           AS ranked,
                SUM(CASE WHEN c.recommendation='Shortlist'
                         THEN 1 ELSE 0 END)                           AS shortlist,
                SUM(CASE WHEN c.recommendation='Maybe'
                         THEN 1 ELSE 0 END)                           AS maybe,
                SUM(CASE WHEN c.recommendation='Reject'
                         THEN 1 ELSE 0 END)                           AS reject,
                MAX(c.ranked_at)                                      AS last_run
            FROM jobs j
            LEFT JOIN candidates c ON c.job_label = j.job_label
            WHERE COALESCE(j.department, 'Uncategorized') = %s
            GROUP BY j.job_label, j.job_title, j.department
            ORDER BY last_run DESC NULLS LAST
        """, (department,))
        rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get("last_run"):
            d["last_run"] = pd.to_datetime(d["last_run"]).strftime("%d %b %Y %H:%M")
        for k in ("total", "ranked", "shortlist", "maybe", "reject"):
            d[k] = int(d.get(k) or 0)
        result.append(d)
    return result


def fetch_candidates_by_department(conn, department: str) -> pd.DataFrame:
    """All ranked candidates in a department, ordered by overall_score desc.

    Adds a sequential 1..N `rank` column after ordering.
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
                c.apply_id, c.candidate_name, c.job_label,
                COALESCE(j.department, 'Uncategorized')     AS department,
                j.job_title,
                c.email, c.mobile, c.location,
                c.degree, c.university, c.experience_detail,
                c.age, c.expected_salary, c.current_salary,
                c.application_date,
                c.bdjobs_score, c.has_uploaded_cv, c.pdf_path,
                c.overall_score, c.skills_score, c.experience_score,
                c.leadership_score, c.education_score, c.culture_fit_score,
                c.experience_years, c.recommendation,
                c.reasoning, c.strengths, c.gaps, c.risk_flags,
                c.hr_override, c.hr_note,
                c.pdf_text_chars, c.jd_used, c.rank_error, c.ranked_at
            FROM candidates c
            JOIN jobs j ON j.job_label = c.job_label
            WHERE COALESCE(j.department, 'Uncategorized') = %s
              AND c.overall_score IS NOT NULL
            ORDER BY c.overall_score DESC NULLS LAST,
                     c.ranked_at DESC NULLS LAST
        """, (department,))
        rows = cur.fetchall()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df.insert(0, "rank", range(1, len(df) + 1))
    df["experience_years"] = pd.to_numeric(df["experience_years"], errors="coerce")
    df["age"]              = pd.to_numeric(df["age"], errors="coerce")
    df["ranked_at"]        = pd.to_datetime(df["ranked_at"]).dt.strftime("%d %b %Y %H:%M")
    return df


def get_job_department(conn, job_label: str) -> str:
    """Small helper — returns department for a job_label (or 'Uncategorized')."""
    with conn.cursor() as cur:
        cur.execute("SELECT department FROM jobs WHERE job_label = %s", (job_label,))
        row = cur.fetchone()
    return (row[0] if row and row[0] else "Uncategorized")


def set_job_department(conn, job_label: str, department: str) -> None:
    """Upsert department for a job_label. Used by the assign-dept UI and ranker."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO jobs (job_label, department)
            VALUES (%s, %s)
            ON CONFLICT (job_label) DO UPDATE SET
                department = EXCLUDED.department,
                updated_at = NOW()
        """, (job_label, department))


def fetch_job(conn, job_label: str) -> dict:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT * FROM jobs WHERE job_label = %s", (job_label,))
        row = cur.fetchone()
    return dict(row) if row else {}


def create_job(job_data: dict):
    conn = fresh_conn()
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO jobs
                (job_label, job_title, department, jd_text, required_skills,
                 red_flags, min_experience, education_req,
                 weight_skills, weight_exp, weight_edu,
                 weight_leadership, weight_culture,
                 interviewer_notes, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pending')
            ON CONFLICT (job_label) DO UPDATE SET
                job_title         = EXCLUDED.job_title,
                department        = EXCLUDED.department,
                jd_text           = EXCLUDED.jd_text,
                required_skills   = EXCLUDED.required_skills,
                red_flags         = EXCLUDED.red_flags,
                min_experience    = EXCLUDED.min_experience,
                education_req     = EXCLUDED.education_req,
                weight_skills     = EXCLUDED.weight_skills,
                weight_exp        = EXCLUDED.weight_exp,
                weight_edu        = EXCLUDED.weight_edu,
                weight_leadership = EXCLUDED.weight_leadership,
                weight_culture    = EXCLUDED.weight_culture,
                interviewer_notes = EXCLUDED.interviewer_notes,
                status            = 'Pending'
        """, (
            job_data["job_label"],       job_data["job_title"],
            job_data["department"],      job_data["jd_text"],
            job_data["required_skills"], job_data["red_flags"],
            job_data["min_experience"],  job_data["education_req"],
            job_data["weight_skills"],   job_data["weight_exp"],
            job_data["weight_edu"],
            job_data.get("weight_leadership", 10),
            job_data.get("weight_culture",    5),
            job_data["interviewer_notes"],
        ))
    conn.close()


def update_job(job_data: dict):
    """Update an existing job's parameters without touching status or timestamps."""
    conn = fresh_conn()
    with conn.cursor() as cur:
        cur.execute("""
            UPDATE jobs SET
                job_title         = %s,
                department        = %s,
                jd_text           = %s,
                required_skills   = %s,
                red_flags         = %s,
                min_experience    = %s,
                education_req     = %s,
                weight_skills     = %s,
                weight_exp        = %s,
                weight_edu        = %s,
                weight_leadership = %s,
                weight_culture    = %s,
                interviewer_notes = %s,
                updated_at        = NOW()
            WHERE job_label = %s
        """, (
            job_data["job_title"],       job_data["department"],
            job_data["jd_text"],         job_data["required_skills"],
            job_data["red_flags"],       job_data["min_experience"],
            job_data["education_req"],   job_data["weight_skills"],
            job_data["weight_exp"],      job_data["weight_edu"],
            job_data.get("weight_leadership", 10),
            job_data.get("weight_culture",    5),
            job_data["interviewer_notes"],
            job_data["job_label"],
        ))
    conn.close()


def update_job_status(job_label: str, status: str):
    conn = fresh_conn()
    with conn.cursor() as cur:
        if status == "Complete":
            cur.execute(
                "UPDATE jobs SET status=%s, last_ranked_at=NOW() WHERE job_label=%s",
                (status, job_label)
            )
        else:
            cur.execute(
                "UPDATE jobs SET status=%s WHERE job_label=%s",
                (status, job_label)
            )
    conn.close()

# ── Candidate queries ──────────────────────────────────────────────────────────

def fetch_candidates(conn, job_label: str) -> pd.DataFrame:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT apply_id, candidate_name, email, mobile, location,
                   degree, university, experience_detail,
                   age, expected_salary, current_salary, application_date,
                   bdjobs_score, has_uploaded_cv, pdf_path,
                   overall_score, skills_score, experience_score,
                   leadership_score, education_score, culture_fit_score,
                   edu_tier_score, edu_degree_score, edu_gpa_score,
                   experience_years, strengths, gaps, risk_flags,
                   recommendation, reasoning, hr_override, hr_note,
                   ranked_at, rank_error, pdf_text_chars
            FROM candidates
            WHERE job_label = %s
            ORDER BY overall_score DESC NULLS LAST
        """, (job_label,))
        rows = cur.fetchall()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["experience_years"] = pd.to_numeric(df["experience_years"], errors="coerce")
    df["age"]              = pd.to_numeric(df["age"], errors="coerce")
    df["ranked_at"]        = pd.to_datetime(df["ranked_at"]).dt.strftime("%d %b %Y %H:%M")
    return df


def delete_candidate(job_label: str, apply_id: str) -> bool:
    """Delete a candidate from the database by job_label + apply_id.
    Also removes any associated audit log entries.
    Returns True if a row was deleted."""
    conn = fresh_conn()
    try:
        with conn.cursor() as cur:
            # Delete audit log entries first (FK-safe)
            try:
                cur.execute(
                    "DELETE FROM hr_audit_log WHERE job_label=%s AND apply_id=%s",
                    (job_label, apply_id),
                )
            except Exception:
                pass  # table may not exist yet
            cur.execute(
                "DELETE FROM candidates WHERE job_label=%s AND apply_id=%s",
                (job_label, apply_id),
            )
            deleted = cur.rowcount > 0
        return deleted
    finally:
        conn.close()


def delete_candidates_bulk(job_label: str, apply_ids: list[str]) -> int:
    """Delete multiple candidates from the database. Returns count deleted."""
    if not apply_ids:
        return 0
    conn = fresh_conn()
    try:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "DELETE FROM hr_audit_log WHERE job_label=%s AND apply_id = ANY(%s)",
                    (job_label, apply_ids),
                )
            except Exception:
                pass  # table may not exist yet
            cur.execute(
                "DELETE FROM candidates WHERE job_label=%s AND apply_id = ANY(%s)",
                (job_label, apply_ids),
            )
            deleted = cur.rowcount
        return deleted
    finally:
        conn.close()


def clear_job_candidates(job_label: str) -> int:
    """Delete ALL candidates (ranked or not) for a job. Also removes audit logs.
    Returns count deleted."""
    conn = fresh_conn()
    try:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "DELETE FROM hr_audit_log WHERE job_label = %s",
                    (job_label,),
                )
            except Exception:
                pass
            cur.execute(
                "DELETE FROM candidates WHERE job_label = %s",
                (job_label,),
            )
            deleted = cur.rowcount
        return deleted
    finally:
        conn.close()


def save_hr_override(job_label: str, apply_id: str, override: str, note: str,
                      hr_user: str = "HR"):
    """Update hr_override on the candidates row AND append a row to
    hr_audit_log (Part 7.5) so the change is traceable for compliance."""
    conn = fresh_conn()
    with conn.cursor() as cur:
        cur.execute(
            "SELECT hr_override FROM candidates WHERE job_label=%s AND apply_id=%s",
            (job_label, apply_id),
        )
        old_row = cur.fetchone()
        old_value = old_row[0] if old_row else None

        cur.execute(
            "UPDATE candidates SET hr_override=%s, hr_note=%s "
            "WHERE job_label=%s AND apply_id=%s",
            (override, note, job_label, apply_id),
        )
        cur.execute(
            """
            INSERT INTO hr_audit_log
                (job_label, apply_id, hr_user, old_value, new_value, note)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (job_label, apply_id, hr_user, old_value, override, note),
        )
    conn.close()


def fetch_audit_log(job_label: str | None = None,
                    apply_id:  str | None = None) -> pd.DataFrame:
    """Return the HR override audit log, optionally filtered by job or candidate."""
    conn = fresh_conn()
    where: list[str] = []
    params: list = []
    if job_label:
        where.append("job_label = %s"); params.append(job_label)
    if apply_id:
        where.append("apply_id = %s");  params.append(apply_id)
    clause = ("WHERE " + " AND ".join(where)) if where else ""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"""
            SELECT id, job_label, apply_id, hr_user, old_value, new_value,
                   note, changed_at
              FROM hr_audit_log
              {clause}
             ORDER BY changed_at DESC
             LIMIT 500
        """, params)
        rows = cur.fetchall()
    conn.close()
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def find_duplicate_applications(apply_id: str | None = None,
                                 email:    str | None = None,
                                 mobile:   str | None = None,
                                 exclude_job: str | None = None) -> list[dict]:
    """Part 7.7 — find other job postings the same person applied to.

    Match by email OR mobile when either is supplied (case-insensitive).
    """
    if not (email or mobile):
        return []
    conn = fresh_conn()
    params: list = []
    clauses: list[str] = []
    if email:
        clauses.append("LOWER(email) = LOWER(%s)"); params.append(email)
    if mobile:
        clauses.append("mobile = %s"); params.append(mobile)
    sql = f"""
        SELECT c.apply_id, c.candidate_name, c.job_label,
               j.job_title, c.recommendation, c.overall_score
          FROM candidates c
          LEFT JOIN jobs j ON j.job_label = c.job_label
         WHERE ({" OR ".join(clauses)})
    """
    if exclude_job:
        sql += " AND c.job_label <> %s"
        params.append(exclude_job)
    sql += " ORDER BY c.overall_score DESC NULLS LAST LIMIT 10"
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows] if rows else []


def associate_candidates_with_job(department: str, new_job_label: str) -> int:
    """Associate all candidates in a department with a specific job_label.

    This is useful when candidates were imported with generic job_labels
    (e.g., 'bdjobs_1463602') but should be associated with a proper
    job from BDJOBS_JOB_REGISTRY (e.g., 'Delivery-Manager').

    Returns the number of candidates updated.
    """
    conn = fresh_conn()
    updated = 0
    with conn.cursor() as cur:
        # First get all distinct job_labels in this department
        cur.execute("""
            SELECT DISTINCT c.job_label
            FROM candidates c
            JOIN jobs j ON j.job_label = c.job_label
            WHERE COALESCE(j.department, 'Uncategorized') = %s
        """, (department,))
        old_labels = [row[0] for row in cur.fetchall()]

        # Update candidates to point to the new job_label
        for old_label in old_labels:
            cur.execute("""
                UPDATE candidates
                SET job_label = %s
                WHERE job_label = %s
            """, (new_job_label, old_label))
            updated += cur.rowcount

        # Ensure the new job exists in jobs table
        cur.execute("""
            INSERT INTO jobs (job_label, department, status)
            VALUES (%s, %s, 'Complete')
            ON CONFLICT (job_label) DO UPDATE SET
                department = EXCLUDED.department,
                status = 'Complete'
        """, (new_job_label, department))

    conn.close()
    return updated


# ── Global stats ───────────────────────────────────────────────────────────────

# ── Excel export ───────────────────────────────────────────────────────────────

def _to_excel_legacy(df: pd.DataFrame, job_label: str) -> bytes:
    """Legacy export — kept for backward compatibility with department view."""
    export = df.copy()
    for col in ["strengths","gaps","risk_flags"]:
        if col in export.columns:
            export[col] = export[col].apply(
                lambda x: "; ".join(x) if isinstance(x, list) else str(x or "")
            )
    for col in ["_idx","pdf_text_chars","rank_error","hr_override","hr_note","has_uploaded_cv"]:
        if col in export.columns:
            export = export.drop(columns=[col])
    rename = {
        "apply_id":"Apply ID","candidate_name":"Candidate Name",
        "overall_score":"AI Overall Score","recommendation":"Recommendation",
        "skills_score":"Skills Score","experience_score":"Experience Score",
        "leadership_score":"Leadership Score","education_score":"Education Score",
        "edu_tier_score":"Edu · Tier","edu_degree_score":"Edu · Degree","edu_gpa_score":"Edu · GPA",
        "culture_fit_score":"Culture Fit Score","experience_years":"Experience (yrs)",
        "strengths":"Strengths","gaps":"Weaknesses","risk_flags":"Risk Flags",
        "reasoning":"AI Reasoning","bdjobs_score":"BDJobs Score (ref)",
        "application_date":"Application Date","age":"Age",
        "expected_salary":"Expected Salary","current_salary":"Current Salary",
        "pdf_path":"Resume File","email":"Email","mobile":"Mobile",
        "location":"Location","degree":"Degree","university":"University",
    }
    export = export.rename(columns={k:v for k,v in rename.items() if k in export.columns})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Rankings")
        ws = writer.sheets["Rankings"]
        from openpyxl.styles import Font, PatternFill, Alignment
        hfill = PatternFill("solid", fgColor="C8102E")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        rec_col = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Recommendation":
                rec_col = i
        rec_colors = {
            "Shortlist": ("DCFCE7","166534"),
            "Maybe":     ("FEF3C7","92400E"),
            "Reject":    ("FEE2E2","991B1B"),
        }
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            bg = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg)
            if rec_col:
                rc = ws.cell(row=row_idx, column=rec_col)
                if str(rc.value) in rec_colors:
                    rbg, rfg = rec_colors[str(rc.value)]
                    rc.fill = PatternFill("solid", fgColor=rbg)
                    rc.font = Font(bold=True, color=rfg)
        for col_cells in ws.columns:
            w = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w+3, 60)
        ws.row_dimensions[1].height = 22
    return buf.getvalue()

# ── Prompt preview ─────────────────────────────────────────────────────────────

def build_prompt_preview(job_data: dict) -> str:
    lines = ["=== AI SCORING PROMPT PREVIEW ===\n"]
    lines.append(f"ROLE: {job_data.get('job_title','Not specified')}")
    if job_data.get("department"):
        lines.append(f"DEPARTMENT: {job_data['department']}")
    if job_data.get("jd_text"):
        jd = job_data["jd_text"]
        lines.append(f"\nJOB DESCRIPTION:\n{jd[:800]}{'...' if len(jd)>800 else ''}")
    if job_data.get("required_skills"):
        lines.append(f"\nREQUIRED SKILLS: {', '.join(job_data['required_skills'])}")
    if job_data.get("min_experience") and job_data["min_experience"] != "Any":
        lines.append(f"MINIMUM EXPERIENCE: {job_data['min_experience']}")
    if job_data.get("education_req") and job_data["education_req"] != "Any":
        lines.append(f"EDUCATION: {job_data['education_req']}")
    ws = job_data.get("weight_skills",     50)
    we = job_data.get("weight_exp",        30)
    wu = job_data.get("weight_edu",        20)
    wl = job_data.get("weight_leadership", 10)
    wc = job_data.get("weight_culture",     5)
    lines.append(
        f"\nSCORING WEIGHTS: Skills {ws}% / Experience {we}% / "
        f"Education {wu}% / Leadership {wl}% / Culture Fit {wc}%"
    )
    if job_data.get("red_flags"):
        lines.append("\nRED FLAGS:\n" + "\n".join(f"  - {f}" for f in job_data["red_flags"]))
    if job_data.get("interviewer_notes"):
        lines.append(f"\nADDITIONAL NOTES:\n{job_data['interviewer_notes']}")
    lines.append("""\n\nOUTPUT FORMAT (strict JSON):
{
  "overall_score": 0-100,
  "skills_score": 0-100,
  "experience_score": 0-100,
  "leadership_score": 0-100,
  "education_score": 0-100,
  "culture_fit_score": 0-100,
  "experience_years": number,
  "strengths": ["..."],
  "gaps": ["..."],
  "recommendation": "Shortlist|Maybe|Reject",
  "reasoning": "2-sentence summary"
}""")
    return "\n".join(lines)

# ── Verdict normalisation ─────────────────────────────────────────────────────

def normalize_verdict(raw: str | None) -> str:
    """
    Collapse LLM or legacy free-text verdicts into the canonical ternary.
    Returns one of: 'Shortlist', 'Maybe', 'Reject'.
    """
    if not raw:
        return "Maybe"
    r = str(raw).lower().strip()
    if r in ("shortlist", "short", "yes", "select", "selected", "pass", "green", "strong fit"):
        return "Shortlist"
    if r in ("reject", "no", "decline", "rejected", "fail", "failed", "red", "weak fit", "not selected"):
        return "Reject"
    return "Maybe"


def fix_inconsistent_verdicts(conn) -> int:
    """
    One-shot DB migration: canonicalise all non-standard recommendation strings.
    Uses a single UPDATE with WHERE … IS DISTINCT FROM guard — safe for production.
    Returns number of rows patched.
    """
    valid = {"Shortlist", "Maybe", "Reject"}
    cur = conn.cursor()
    # Find candidates whose recommendation is not in the canonical set
    cur.execute("""
        SELECT apply_id, recommendation
        FROM candidates
        WHERE recommendation IS NOT NULL
          AND recommendation NOT IN ('Shortlist', 'Maybe', 'Reject')
    """)
    rows = cur.fetchall()
    patched = 0
    for apply_id, raw in rows:
        canonical = normalize_verdict(raw)
        if canonical != raw:
            cur.execute("""
                UPDATE candidates
                SET recommendation = %s
                WHERE apply_id = %s AND recommendation IS DISTINCT FROM %s
            """, (canonical, apply_id, canonical))
            patched += cur.rowcount
    conn.commit()
    cur.close()
    return patched


# ── Theme ──────────────────────────────────────────────────────────────────────

def init_theme():
    st.session_state["day_mode"] = True

def theme_toggle():
    st.session_state["day_mode"] = True

def get_css() -> str:
    return CSS

def t(day_val, night_val=None):
    return day_val

# ── Sidebar ────────────────────────────────────────────────────────────────────

def render_sidebar():
    """
    Shared sidebar renderer for all pages.
    Enforces: single red column, no Streamlit default nav, fixed 288px width.
    """
    # ── Inline CSS: must appear INSIDE the sidebar context so it applies ─────
    # The stSidebarNav hide rule is duplicated here as a safety net for pages
    # that load before the page-level CSS block is rendered.
    with st.sidebar:
        st.markdown("""
            <style>
            /* Kill Streamlit's auto-generated folder nav — every page */
            [data-testid="stSidebarNav"]               { display: none !important; }
            [data-testid="stSidebarNavItems"]           { display: none !important; }
            [data-testid="stSidebarNavSeparator"]       { display: none !important; }

            /* Lock sidebar width — never collapsible */
            [data-testid="collapsedControl"]            { display: none !important; }
            [data-testid="stSidebarCollapseButton"]     { display: none !important; }
            [data-testid="stSidebar"] {
                min-width: 288px !important;
                max-width: 288px !important;
                transform: none !important;
            }
            [data-testid="stSidebar"][aria-expanded="false"] {
                min-width: 288px !important;
                transform: none !important;
                margin-left: 0 !important;
            }

            /* Red background — full sidebar */
            [data-testid="stSidebar"]       { background-color: #C8102E !important; border-right: none !important; }
            [data-testid="stSidebar"] > div { background-color: #C8102E !important; }

            /* All sidebar text white */
            [data-testid="stSidebar"] *     { color: #FFFFFF !important; font-family: 'Inter', sans-serif !important; font-size: 14px !important; }

            /* Nav link hover */
            [data-testid="stSidebar"] a:hover { background: rgba(255,255,255,0.12) !important; border-radius: 6px; }

            /* Active page link highlight */
            [data-testid="stSidebar"] [aria-current="page"]         { background: rgba(255,255,255,0.18) !important; border-radius: 6px; }
            [data-testid="stSidebar"] [aria-current="page"] a,
            [data-testid="stSidebar"] [aria-current="page"] a p     { font-weight: 700 !important; }

            /* Section label style */
            [data-testid="stSidebar"] .nav-label {
                color: rgba(255,255,255,0.65) !important;
                font-size: 0.65rem !important;
                text-transform: uppercase;
                letter-spacing: 0.1em;
                margin: 1.5rem 0 0.5rem 0;
                display: block;
                padding-left: 0.3rem;
            }

            /* Divider */
            [data-testid="stSidebar"] hr.divider {
                border: none;
                border-top: 1px solid rgba(255,255,255,0.2) !important;
                margin: 0.8rem 0;
            }
            </style>
        """, unsafe_allow_html=True)

        # ── Logo or text brand ─────────────────────────────────────────────────
        logo_path = LOGO_PATH
        if os.path.exists(logo_path):
            st.markdown('<div style="padding:2% 2% 1rem 2%;">', unsafe_allow_html=True)
            st.image(logo_path, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown("""
                <div style="padding:0.6rem 0 1.4rem;">
                    <div style="font-size:1.05rem;font-weight:700;color:#FFFFFF !important;">
                        Olympic Industries PLC
                    </div>
                    <div style="font-size:0.72rem;color:rgba(255,255,255,0.7) !important;
                                margin-top:3px;text-transform:uppercase;letter-spacing:0.1em;">
                        HR Intelligence Platform
                    </div>
                </div>
            """, unsafe_allow_html=True)

        # ── Auth section ─────────────────────────────────────────────────────────
        user = st.session_state.get("user")
        if user:
            st.markdown('<hr class="divider">', unsafe_allow_html=True)
            st.markdown(
                f"""
                <div style="padding:0.3rem 0.4rem;">
                    <div style="font-size:0.82rem;font-weight:600;color:#FFFFFF !important;">
                        👤 {user.get('display_name', user['username'])}
                    </div>
                    <div style="font-size:0.65rem;color:rgba(255,255,255,0.65) !important;">
                        {'🔑 Admin' if user.get('role') == 'admin' else '👤 User'}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.button("🚪  Logout", use_container_width=True, type="secondary"):
                try:
                    _conn = fresh_conn()
                    log_audit(_conn, user.get("id"), user.get("username"), "LOGOUT")
                except Exception:
                    pass
                st.session_state.pop("user", None)
                st.session_state.clear()
                st.rerun()

        # ── Navigation (only when logged in) ─────────────────────────────────────
        st.markdown('<div class="nav-label">Navigation</div>', unsafe_allow_html=True)

        def _safe_page_link(page: str, label: str) -> None:
            """Wrap st.page_link in try-except for Streamlit Cloud compatibility.
            Falls back to a markdown link if page_link raises (e.g. KeyError on url_pathname)."""
            try:
                st.page_link(page, label=label)
            except Exception:
                if page == "Home.py":
                    href = "/"
                else:
                    name = page.replace("pages/", "").replace(".py", "")
                    name = re.sub(r"^\d+_", "", name)
                    href = f"/{name}"
                st.markdown(
                    f"<a href='{href}' style='display:block;padding:0.35rem 0.6rem;"
                    f"color:#FFFFFF;text-decoration:none;border-radius:6px;"
                    f"font-size:0.9rem;'>"
                    f"{label}</a>",
                    unsafe_allow_html=True,
                )

        if user:
            _safe_page_link("Home.py",                           label="📋  Dashboard")
            _safe_page_link("pages/0_Download_CVs.py",           label="⬇️⬆️ Download/Upload CVs")
            _safe_page_link("pages/1_Department_Rankings.py",    label="🏢  Department Rankings")
            _safe_page_link("pages/2_Job_Rankings.py",           label="📊  Job Rankings")
            _safe_page_link("pages/3_New_Job.py",                label="📝  New Job Posting")
            _safe_page_link("pages/4_Processing_Status.py",      label="⏳  Processing Status")
            _safe_page_link("pages/6_Compare_Candidates.py",     label="⚖️  Compare Candidates")
            _safe_page_link("pages/5_Settings.py",               label="⚙️  Settings")
            if user.get("role") == "admin":
                _safe_page_link("pages/7_Admin.py", label="🔐  Admin Panel")
        else:
            _safe_page_link("pages/0_Login.py", label="🔐  Login")

def safe_switch_page(page: str) -> None:
    """Cloud-safe wrapper around st.switch_page.
    Streamlit Cloud may fail to resolve page paths, so we fall back
    to st.page_link redirect or markdown-based redirect."""
    try:
        st.switch_page(page)
    except Exception:
        # Fallback: use st.page_link as redirect trigger
        try:
            st.page_link(page)
        except Exception:
            name = page.replace("pages/", "").replace(".py", "")
            name = re.sub(r"^\d+_", "", name)
            st.markdown(
                f'<meta http-equiv="refresh" content="0;URL=/{name}">',
                unsafe_allow_html=True,
            )
            st.stop()


def fetch_global_stats(conn=None) -> dict:
    """Global candidate stats. `conn` is accepted for API symmetry but we
    always open a fresh autocommit connection so Streamlit cache can't serve
    a stale snapshot."""
    fresh = fresh_conn()
    with fresh.cursor() as cur:
        # Only count jobs that actually have candidates (clean-interface rule)
        cur.execute("""
            SELECT
                (SELECT COUNT(DISTINCT job_label) FROM candidates) AS total_jobs,
                COUNT(*)                                            AS total_candidates,
                SUM(CASE WHEN overall_score IS NULL THEN 1 ELSE 0 END) AS pending,
                ROUND(AVG(overall_score))                           AS avg_score
            FROM candidates
        """)
        row = cur.fetchone()
    fresh.close()
    return {
        "total_jobs":       row[0] or 0,
        "total_candidates": row[1] or 0,
        "pending":          row[2] or 0,
        "avg_score":        row[3] or 0,
    }

# ── CSS ────────────────────────────────────────────────────────────────────────

def _is_streamlit_cloud() -> bool:
    """True when running on Streamlit Community Cloud (no GUI, no local Ollama/Playwright)."""
    import platform
    # Check 1: explicit Streamlit Cloud env var
    if os.environ.get("STREAMLIT_SHARING", "").lower() == "true":
        return True
    # Check 2: Streamlit server port is set (always present on Cloud)
    if os.environ.get("STREAMLIT_SERVER_PORT"):
        return True
    # Check 3: Linux with no DISPLAY (headless server)
    if platform.system() == "Linux" and os.environ.get("DISPLAY") is None:
        return True
    # Check 4: Cloud mounts repos under /mount/src
    if Path("/mount/src").exists():
        return True
    return False

CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

#MainMenu, footer, header { visibility: hidden !important; }
[data-testid="collapsedControl"]        { display: none !important; }
[data-testid="stSidebarCollapseButton"] { display: none !important; }
[data-testid="stSidebar"] {
    min-width: 288px !important;
    max-width: 288px !important;
    transform: none !important;
}
[data-testid="stSidebar"][aria-expanded="false"] {
    min-width: 288px !important;
    transform: none !important;
    margin-left: 0 !important;
}

[data-testid="stSidebarNav"] { display: none !important; }

.block-container { padding: 1.8rem 2.2rem 3rem !important; max-width: 1440px; }

html, body, p, span, div, label, li, td, th, input, textarea, select {
    font-family: 'Inter', sans-serif !important;
    font-size: 15px;
}

[data-testid="stSidebar"] { background-color: #C8102E !important; border-right: none !important; }
[data-testid="stSidebar"] > div { background-color: #C8102E !important; }
[data-testid="stSidebar"] * { color: #FFFFFF !important; font-family: 'Inter', sans-serif !important; font-size: 14px !important; }
[data-testid="stSidebar"] a { color: #FFFFFF !important; text-decoration: none !important; }
[data-testid="stSidebar"] a:hover { color: rgba(255,255,255,0.75) !important; }
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.12) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    color: #FFFFFF !important;
    font-size: 0.83rem !important;
    width: 100% !important;
    border-radius: 6px !important;
    padding: 0.45rem 0.75rem !important;
    margin-bottom: 3px !important;
    text-align: left !important;
    font-weight: 400 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.22) !important;
    border-color: rgba(255,255,255,0.4) !important;
}

.nav-label {
    font-size: 0.68rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: rgba(255,255,255,0.6) !important;
    margin-bottom: 0.5rem;
    padding-bottom: 0.3rem;
    border-bottom: 1px solid rgba(255,255,255,0.2);
}

[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background-color: rgba(255,255,255,0.2) !important;
    border-radius: 4px !important;
}
[data-testid="stSidebar"] span[data-baseweb="tag"] * {
    color: #FFFFFF !important;
}

html, body { background-color: #F8FAFC !important; }
.stApp, section[data-testid="stMain"], section[data-testid="stMain"] > div,
.block-container { background-color: #F8FAFC !important; }

html, body, p, span, label, li, td, th { color: #1E293B !important; }

.page-title { font-size: 1.55rem; font-weight: 700; line-height: 1.2; margin-bottom: 0; color: #1E293B !important; }
.page-sub   { font-size: 0.85rem; margin-top: 3px; color: #64748B !important; }
.section-hd {
    font-size: 0.72rem; font-weight: 600; letter-spacing: 0.14em;
    text-transform: uppercase; margin-bottom: 0.75rem;
    padding-bottom: 0.45rem; color: #64748B !important;
    border-bottom: 1px solid #E2E8F0;
}
.divider { border: none; border-top: 1px solid #E2E8F0; margin: 1.2rem 0; }

.metric-card {
    border-radius: 10px; padding: 1.1rem 1.3rem;
    background: #FFFFFF; border: 1px solid #E2E8F0;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.metric-val { font-size: 2.1rem; font-weight: 700; line-height: 1; color: #1E293B !important; }
.metric-lbl { font-size: 0.74rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 4px; color: #64748B !important; }

.verdict-badge     { display: inline-block; padding: 3px 11px; border-radius: 12px; font-size: 0.78rem; font-weight: 600; }
.verdict-shortlist { background: #DCFCE7; color: #166534 !important; }
.verdict-maybe     { background: #FEF3C7; color: #92400E !important; }
.verdict-reject    { background: #FEE2E2; color: #991B1B !important; }

.status-badge      { display: inline-block; padding: 2px 9px; border-radius: 10px; font-size: 0.73rem; font-weight: 500; }
.status-complete   { background: #DCFCE7; color: #166534 !important; }
.status-processing { background: #DBEAFE; color: #1E40AF !important; }
.status-pending    { background: #F1F5F9; color: #475569 !important; }
.status-error      { background: #FEE2E2; color: #991B1B !important; }

.score-pill   { display: inline-block; border-radius: 6px; padding: 4px 12px; font-size: 0.8rem; font-weight: 600; margin: 3px 4px 3px 0; }
.score-green  { background: #DCFCE7; color: #166534 !important; }
.score-yellow { background: #FEF3C7; color: #92400E !important; }
.score-red    { background: #FEE2E2; color: #991B1B !important; }

.flag-chip { display: inline-block; border-radius: 4px; padding: 2px 8px; font-size: 0.73rem; font-weight: 500; margin: 2px 3px 2px 0; background: #FEE2E2; color: #991B1B !important; }

.cand-name-lg { font-size: 1.4rem; font-weight: 700; line-height: 1.2; }
.cand-meta-sm { font-size: 0.83rem; margin-top: 3px; }
.hint-text    { font-size: 0.8rem; color: #94A3B8 !important; }

div[data-testid="stDataFrame"] { border: 1px solid #E2E8F0 !important; border-radius: 8px !important; box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important; }
div[data-testid="stDataFrame"] thead th {
    background-color: #C8102E !important;
    color: #FFFFFF !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.03em !important;
}
div[data-testid="stDataFrame"] { color-scheme: light !important; }

div[data-baseweb="select"] > div { background-color: #FFFFFF !important; border-color: #CBD5E1 !important; border-radius: 6px !important; }
div[data-baseweb="select"] * { color: #1E293B !important; }
div[data-baseweb="popover"] { background-color: #FFFFFF !important; border: 1px solid #E2E8F0 !important; border-radius: 8px !important; }
div[data-baseweb="popover"] * { background-color: #FFFFFF !important; color: #1E293B !important; }

# _________Sidebar Tag Overrides (for better visibility against red background)_________
[data-testid="stSidebar"] span[data-baseweb="tag"] { background-color: rgba(0,0,0,0.25) !important; border-radius: 4px !important; }
[data-testid="stSidebar"] span[data-baseweb="tag"] * { color: #6E1400 !important; }

span[data-baseweb="tag"] { background-color: #FECDD3 !important; border-radius: 4px !important; }
span[data-baseweb="tag"] * { color: #9F1239 !important; }

textarea, input[type="text"] { background-color: #FFFFFF !important; border: 1px solid #CBD5E1 !important; border-radius: 6px !important; color: #1E293B !important; }

button[kind="primary"], button[data-testid="baseButton-primary"] { background: #C8102E !important; color: #FFFFFF !important; border: none !important; border-radius: 6px !important; font-family: 'Inter', sans-serif !important; font-size: 0.85rem !important; font-weight: 500 !important; }
button[kind="primary"] p, button[data-testid="baseButton-primary"] p { color: #FFFFFF !important; }
button[kind="primary"] span, button[data-testid="baseButton-primary"] span { color: #FFFFFF !important; }
button[kind="primary"]:hover, button[data-testid="baseButton-primary"]:hover { background: #A00D24 !important; }
button[kind="secondary"], button[data-testid="baseButton-secondary"] { background: #FFFFFF !important; color: #C8102E !important; border: 1px solid #E2E8F0 !important; border-radius: 6px !important; font-family: 'Inter', sans-serif !important; font-size: 0.85rem !important; font-weight: 500 !important; }
button[kind="secondary"]:hover, button[data-testid="baseButton-secondary"]:hover { background: #FFF1F2 !important; border-color: #C8102E !important; color: #A00D24 !important; }
.stDownloadButton > button { background: #FFFFFF !important; border: 1px solid #CBD5E1 !important; color: #C8102E !important; border-radius: 6px !important; font-family: 'Inter', sans-serif !important; font-size: 0.85rem !important; font-weight: 500 !important; }
.stDownloadButton > button:hover { border-color: #C8102E !important; background: #FFF1F2 !important; }

div[data-testid="stAlert"] { border-radius: 8px !important; font-size: 0.88rem !important; }
.stMarkdown p, .stMarkdown li { color: #374151 !important; font-size: 0.92rem; line-height: 1.65; }

/* ── Department cards (Department Rankings page + Home grouped view) ── */
.dept-card {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-left: 4px solid #C8102E;
    border-radius: 10px;
    padding: 1.2rem 1.4rem;
    margin-bottom: 0.7rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.dept-name {
    font-size: 1.05rem;
    font-weight: 700;
    color: #1E293B !important;
    margin-bottom: 6px;
    letter-spacing: 0.01em;
}
.dept-stats {
    font-size: 0.85rem;
    color: #475569 !important;
}
.dept-meta {
    font-size: 0.72rem;
    color: #94A3B8 !important;
    margin-top: 4px;
    letter-spacing: 0.02em;
}
.dept-group-title {
    font-size: 0.78rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #C8102E !important;
    padding: 0.8rem 0 0.3rem;
    border-bottom: 1px solid #E2E8F0;
    margin-bottom: 0.5rem;
}
.rank-pill {
    display: inline-block;
    background: #C8102E;
    color: #FFFFFF !important;
    border-radius: 10px;
    padding: 2px 10px;
    font-size: 0.76rem;
    font-weight: 700;
    letter-spacing: 0.02em;
}

/* Preserve Material Symbols icon font everywhere (main + sidebar). The
   global `span` and `[data-testid="stSidebar"] *` font-family overrides
   above would otherwise render expander chevrons as literal ligature
   text such as "keyboard_arrow_down". Placed last so cascade wins. */
[data-testid="stIconMaterial"],
[data-testid="stIconMaterial"] *,
[data-testid="stSidebar"] [data-testid="stIconMaterial"],
[data-testid="stSidebar"] [data-testid="stIconMaterial"] * {
    font-family: 'Material Symbols Rounded', 'Material Icons' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}
</style>"""

CUSTOM_CSS = CSS


# ═══════════════════════════════════════════════════════════════════════════════
# BDJobs credential helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _encode_pwd(pwd: str) -> str:
    """Basic obfuscation — not encryption, but prevents casual shoulder-surfing."""
    import base64
    return base64.b64encode(pwd.encode("utf-8")).decode("ascii")


def _decode_pwd(enc: str) -> str:
    import base64
    return base64.b64decode(enc.encode("ascii")).decode("utf-8")


def save_bdjobs_credentials(conn, username: str, password: str) -> None:
    """Upsert BDJobs credentials."""
    enc = _encode_pwd(password)
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO bdjobs_credentials (id, username, password, updated_at)
            VALUES (1, %s, %s, NOW())
            ON CONFLICT (id) DO UPDATE SET
                username = EXCLUDED.username,
                password = EXCLUDED.password,
                updated_at = NOW()
        """, (username, enc))


def get_bdjobs_credentials(conn) -> dict | None:
    """Return {"username": str, "password": str} or None."""
    with conn.cursor() as cur:
        cur.execute("SELECT username, password FROM bdjobs_credentials ORDER BY id LIMIT 1")
        row = cur.fetchone()
    if row:
        return {"username": row[0], "password": _decode_pwd(row[1])}
    return None


def has_bdjobs_credentials(conn) -> bool:
    """Return True if credentials are stored."""
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM bdjobs_credentials LIMIT 1")
        return cur.fetchone() is not None


def _build_detail_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Generate skills_details, experience_details, education_details,
    and weaknesses columns from existing candidate data for Excel export."""
    export = df.copy()

    # skills_details: combine strengths (skills-focused) + skills_score context
    def _skills_detail(row):
        parts = []
        strengths = row.get("strengths")
        if isinstance(strengths, list):
            parts = [s for s in strengths if s]
        elif isinstance(strengths, str) and strengths.strip() and strengths.strip() != "—":
            parts = [strengths.strip()]
        score = row.get("skills_score")
        if score is not None and int(score or 0) > 0:
            parts.append(f"Skills Score: {int(score)}/100")
        return "; ".join(parts) if parts else ""
    export["skills_details"] = export.apply(_skills_detail, axis=1)

    # experience_details: use experience_detail field + experience_years + experience_score
    def _exp_detail(row):
        parts = []
        exp = row.get("experience_detail") or ""
        if exp:
            entries = [e.strip().replace("*", "").replace("##", " — ") for e in str(exp).split("|") if e.strip()]
            parts.extend(entries)
        years = row.get("experience_years")
        if years and float(years or 0) > 0:
            parts.append(f"Total: {float(years):.1f} years")
        score = row.get("experience_score")
        if score is not None and int(score or 0) > 0:
            parts.append(f"Experience Score: {int(score)}/100")
        return "; ".join(parts) if parts else ""
    export["experience_details"] = export.apply(_exp_detail, axis=1)

    # education_details: degree + university + sub-scores
    def _edu_detail(row):
        parts = []
        degree = str(row.get("degree") or "")
        univ = str(row.get("university") or "")
        if degree and degree.lower() != "nan":
            parts.append(degree)
        if univ and univ.lower() != "nan":
            parts.append(univ)
        edu_score = row.get("education_score")
        if edu_score is not None and int(edu_score or 0) > 0:
            parts.append(f"Education Score: {int(edu_score)}/100")
        tier = row.get("edu_tier_score")
        deg_sc = row.get("edu_degree_score")
        gpa = row.get("edu_gpa_score")
        if any(v is not None and int(v or 0) > 0 for v in [tier, deg_sc, gpa]):
            parts.append(f"Tier: {int(tier or 0)}, Degree: {int(deg_sc or 0)}, GPA: {int(gpa or 0)}")
        return "; ".join(parts) if parts else ""
    export["education_details"] = export.apply(_edu_detail, axis=1)

    # weaknesses: from gaps field
    def _weaknesses(row):
        gaps = row.get("gaps")
        if isinstance(gaps, list):
            return "; ".join(g for g in gaps if g)
        elif isinstance(gaps, str) and gaps.strip() and gaps.strip() != "—":
            return gaps.strip()
        return ""
    export["weaknesses"] = export.apply(_weaknesses, axis=1)

    # Convert list columns to strings for Excel
    for col in ["strengths", "risk_flags"]:
        if col in export.columns:
            export[col] = export[col].apply(
                lambda x: "; ".join(x) if isinstance(x, list) else str(x or "")
            )

    return export


def to_excel(
    ranked_df: pd.DataFrame,
    job_label: str,
    unranked_df: pd.DataFrame | None = None,
) -> io.BytesIO:
    """
    Export job candidates to Excel with two sheets:
      1. "Rankings" — all ranked candidates (filters ignored)
      2. "Unranked" — candidates not yet scored (if provided)
    Returns BytesIO buffer for st.download_button.
    """
    # Column order as per prompt requirements
    rank_cols = [
        "apply_id",
        "candidate_name",
        "email",
        "mobile",
        "location",
        "degree",
        "university",
        "experience_detail",
        "age",
        "expected_salary",
        "current_salary",
        "application_date",
        "bdjobs_score",
        "overall_score",
        "recommendation",
        "skills_score",
        "experience_score",
        "education_score",
        "skills_details",
        "experience_details",
        "education_details",
        "strengths",
        "weaknesses",
        "risk_flags",
        "reasoning",
    ]
    unrank_cols = [
        "apply_id",
        "candidate_name",
        "email",
        "mobile",
        "location",
        "degree",
        "university",
        "experience_detail",
        "age",
        "expected_salary",
        "current_salary",
        "application_date",
        "bdjobs_score",
    ]

    # Ensure columns exist (fill missing with empty string)
    def _ensure_cols(df, cols):
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        return df[cols]

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        # ── Sheet 1: Rankings ────────────────────────────────────────────────
        if ranked_df is not None and not ranked_df.empty:
            enriched = _build_detail_columns(ranked_df)
            rdf = _ensure_cols(enriched, rank_cols)
        else:
            rdf = pd.DataFrame(columns=rank_cols)
        rdf.to_excel(writer, index=False, sheet_name="Rankings")

        # ── Sheet 2: Unranked ──────────────────────────────────────────────
        if unranked_df is not None and not unranked_df.empty:
            udf = _ensure_cols(unranked_df.copy(), unrank_cols)
            udf.to_excel(writer, index=False, sheet_name="Unranked")

    out.seek(0)
    return out


def _style_sheet(ws, df: pd.DataFrame, is_ranked: bool = True):
    """
    Apply Olympic red header + conditional formatting (green/amber/red).
    Auto-size columns and freeze header row.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Olympic red header
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Green / Amber / Red fills for verdicts
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            # Align numbers right, text left
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Conditional formatting: recommendation column
    if is_ranked and "recommendation" in df.columns:
        rec_col = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == "recommendation":
                rec_col = get_column_letter(idx)
                break
        if rec_col:
            for row_num in range(2, ws.max_row + 1):
                cell = ws[f"{rec_col}{row_num}"]
                val = str(cell.value or "").strip()
                if val == "Shortlist":
                    cell.fill = green_fill
                elif val == "Maybe":
                    cell.fill = amber_fill
                elif val == "Reject":
                    cell.fill = red_fill

    # Conditional formatting: overall_score column
    if is_ranked and "overall_score" in df.columns:
        score_col = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == "overall_score":
                score_col = get_column_letter(idx)
                break
        if score_col:
            for row_num in range(2, ws.max_row + 1):
                cell = ws[f"{score_col}{row_num}"]
                try:
                    val = float(cell.value or 0)
                    if val >= 70:
                        cell.fill = green_fill
                    elif val >= 50:
                        cell.fill = amber_fill
                    else:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    pass

    # Auto-size columns
    for idx, col in enumerate(df.columns, 1):
        max_len = len(str(col))
        for val in df[col]:
            max_len = max(max_len, len(str(val)[:50]))  # cap at 50 chars
        adjusted = min(max_len + 2, 40)  # cap width at 40
        ws.column_dimensions[get_column_letter(idx)].width = adjusted

    # Freeze header row
    ws.freeze_panes = "A2"

