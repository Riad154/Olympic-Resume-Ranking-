"""
pages/6_Compare_Candidates.py — FEAT-01

Side-by-side comparison of 2-3 candidates ranked under the same job.
Drives off st.session_state["compare_ids"] + st.session_state["compare_job"]
which are set by the Job Rankings page multiselect.
"""

import io
import streamlit as st
import pandas as pd

from db import (
    get_conn, fetch_candidates, get_css, init_theme, render_sidebar, safe_switch_page,
    SCORE_DIMS, VERDICT_CFG, _build_detail_columns,
)

st.set_page_config(
    page_title="Compare Candidates — HR Intelligence",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

init_theme()
st.markdown(get_css(), unsafe_allow_html=True)
render_sidebar()

if not st.session_state.get("user"):
    st.warning("🔒 Please log in to access this page.")
    safe_switch_page("pages/0_Login.py")
    st.stop()


is_day   = st.session_state.get("day_mode", True)
txt_col  = "#1E293B" if is_day else "#E2E8F0"
sub_col  = "#64748B"
card_bg  = "#FFFFFF" if is_day else "#1E2435"
card_bdr = "#E2E8F0" if is_day else "#2D3748"

st.markdown(
    f'<div class="page-title" style="color:{txt_col} !important;">Compare Candidates</div>',
    unsafe_allow_html=True,
)
st.markdown(
    '<div class="page-sub">Side-by-side breakdown of selected candidates</div>',
    unsafe_allow_html=True,
)
st.markdown(f'<hr class="divider" style="border-top:1px solid {card_bdr}">',
            unsafe_allow_html=True)

# ── Session state -----------------------------------------------------------------
compare_ids = st.session_state.get("compare_ids") or []
compare_job = st.session_state.get("compare_job") or ""

if not compare_job or len(compare_ids) < 2:
    st.info(
        "No candidates queued for comparison. Open **Job Rankings**, expand "
        "**“Compare candidates side-by-side”**, pick 2–3 candidates, then "
        "click *Open Comparison*.",
    )
    if st.button("← Back to Job Rankings"):
        safe_switch_page("pages/2_Job_Rankings.py")
    st.stop()

# ── Load -------------------------------------------------------------------------
try:
    conn = get_conn()
except Exception as e:
    st.error(f"Database connection failed: {e}")
    st.stop()

df_all = fetch_candidates(conn, compare_job)
if df_all.empty:
    st.warning(f"No candidates found for **{compare_job}**.")
    st.stop()

df = df_all[df_all["apply_id"].isin(compare_ids)].copy()
if df.empty:
    st.warning("Selected candidates are no longer in the database.")
    st.stop()

# Preserve the user's chosen ordering
df["_order"] = df["apply_id"].apply(lambda a: compare_ids.index(a) if a in compare_ids else 99)
df = df.sort_values("_order").drop(columns="_order").reset_index(drop=True)

# ── Header card row ---------------------------------------------------------------
st.markdown(
    f'<div style="font-size:0.85rem;color:{sub_col};margin-bottom:0.6rem;">'
    f'Job: <b>{compare_job}</b> · {len(df)} candidate(s)</div>',
    unsafe_allow_html=True,
)

cols = st.columns(len(df))
for i, (_, c) in enumerate(df.iterrows()):
    verdict  = str(c.get("hr_override") or c.get("recommendation") or "")
    vcfg     = VERDICT_CFG.get(verdict, {"color": "#64748B", "bg": "#F1F5F9", "icon": "◆"})
    score    = int(c.get("overall_score") or 0)
    name     = str(c.get("candidate_name") or "Unknown")
    apply_id = str(c.get("apply_id") or "")

    with cols[i]:
        st.markdown(f"""
            <div style="background:{card_bg};border:1px solid {card_bdr};
                        border-radius:12px;padding:1.1rem 1rem;margin-bottom:0.8rem;
                        text-align:center;">
                <div style="font-size:1.05rem;font-weight:700;color:{txt_col} !important;">
                    {name}
                </div>
                <div style="font-size:0.74rem;color:{sub_col} !important;margin-top:2px;">
                    Apply ID {apply_id}
                </div>
                <div style="font-size:2.4rem;font-weight:700;color:{txt_col} !important;
                            margin-top:0.6rem;line-height:1;">{score}</div>
                <div style="font-size:0.72rem;color:{sub_col} !important;text-transform:uppercase;
                            letter-spacing:0.08em;margin-top:2px;">Overall</div>
                <div style="margin-top:0.7rem;display:inline-block;
                            background:{vcfg.get('bg','#F1F5F9')};
                            color:{vcfg.get('color','#64748B')} !important;
                            font-weight:600;font-size:0.78rem;
                            padding:3px 12px;border-radius:14px;">
                    {vcfg.get('icon','◆')} {verdict or '—'}
                </div>
            </div>
        """, unsafe_allow_html=True)

# ── Dimension comparison table ----------------------------------------------------
st.markdown(
    f'<div class="section-hd" style="color:{sub_col} !important;'
    f'border-bottom:1px solid {card_bdr};">Score Dimensions</div>',
    unsafe_allow_html=True,
)

dim_rows = []
for key, label in SCORE_DIMS:
    row = {"Dimension": label}
    for _, c in df.iterrows():
        col_name = str(c.get("candidate_name") or c.get("apply_id"))
        row[col_name] = int(c.get(key) or 0) if c.get(key) is not None else None
    dim_rows.append(row)

# Education sub-scores (FEAT-05 surfaces these in detail; reuse here)
for sub_key, sub_label in [
    ("edu_tier_score",   "  └ University Tier (50%)"),
    ("edu_degree_score", "  └ Degree Level (30%)"),
    ("edu_gpa_score",    "  └ GPA / Result (20%)"),
]:
    if sub_key in df.columns:
        row = {"Dimension": sub_label}
        for _, c in df.iterrows():
            col_name = str(c.get("candidate_name") or c.get("apply_id"))
            v = c.get(sub_key)
            row[col_name] = int(v) if v is not None else None
        dim_rows.append(row)

dim_df = pd.DataFrame(dim_rows)
candidate_cols = [col for col in dim_df.columns if col != "Dimension"]
column_config = {"Dimension": st.column_config.TextColumn("Dimension", width="medium")}
for col in candidate_cols:
    column_config[col] = st.column_config.ProgressColumn(
        col, min_value=0, max_value=100, format="%d",
    )

st.dataframe(
    dim_df, use_container_width=True, hide_index=True,
    column_config=column_config,
)

# ── Strengths / Gaps / Risk flags / Reasoning ------------------------------------
st.markdown(
    f'<div class="section-hd" style="color:{sub_col} !important;'
    f'border-bottom:1px solid {card_bdr};margin-top:0.6rem;">Qualitative Notes</div>',
    unsafe_allow_html=True,
)

note_cols = st.columns(len(df))
for i, (_, c) in enumerate(df.iterrows()):
    name      = str(c.get("candidate_name") or c.get("apply_id"))
    strengths = str(c.get("strengths") or "—")
    gaps      = str(c.get("gaps") or "—")
    flags     = str(c.get("risk_flags") or "—")
    reasoning = str(c.get("reasoning") or "—")
    with note_cols[i]:
        st.markdown(f"""
            <div style="background:{card_bg};border:1px solid {card_bdr};
                        border-radius:10px;padding:1rem;margin-bottom:0.6rem;">
                <div style="font-weight:600;color:{txt_col} !important;
                            font-size:0.92rem;margin-bottom:0.6rem;">{name}</div>
                <div style="font-size:0.75rem;color:{sub_col};text-transform:uppercase;
                            letter-spacing:0.08em;margin-top:0.6rem;">Strengths</div>
                <div style="font-size:0.85rem;color:{txt_col} !important;line-height:1.55;
                            margin-top:0.2rem;">{strengths}</div>
                <div style="font-size:0.75rem;color:{sub_col};text-transform:uppercase;
                            letter-spacing:0.08em;margin-top:0.7rem;">Gaps</div>
                <div style="font-size:0.85rem;color:{txt_col} !important;line-height:1.55;
                            margin-top:0.2rem;">{gaps}</div>
                <div style="font-size:0.75rem;color:{sub_col};text-transform:uppercase;
                            letter-spacing:0.08em;margin-top:0.7rem;">Risk Flags</div>
                <div style="font-size:0.85rem;color:{txt_col} !important;line-height:1.55;
                            margin-top:0.2rem;">{flags}</div>
                <div style="font-size:0.75rem;color:{sub_col};text-transform:uppercase;
                            letter-spacing:0.08em;margin-top:0.7rem;">AI Reasoning</div>
                <div style="font-size:0.83rem;color:{txt_col} !important;line-height:1.55;
                            margin-top:0.2rem;">{reasoning}</div>
            </div>
        """, unsafe_allow_html=True)

# ── Export comparison to Excel ---------------------------------------------------
st.markdown(f'<hr class="divider" style="border-top:1px solid {card_bdr}">',
            unsafe_allow_html=True)

def _compare_to_excel(compare_df: pd.DataFrame, job: str) -> bytes:
    """Export comparison candidates to Excel with detail columns."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    export = _build_detail_columns(compare_df)

    # Select and order columns for comparison export
    export_cols = [
        "candidate_name", "overall_score", "recommendation",
        "skills_score", "experience_score", "leadership_score",
        "education_score", "culture_fit_score",
        "skills_details", "experience_details", "education_details",
        "strengths", "weaknesses", "risk_flags", "reasoning",
        "experience_years", "degree", "university", "location",
    ]
    for c in export_cols:
        if c not in export.columns:
            export[c] = ""
    export = export[export_cols]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Comparison")
        ws = writer.sheets["Comparison"]
        header_fill = PatternFill("solid", fgColor="C8102E")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
        for idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(idx)
            max_len = max(len(str(ws.cell(row=r, column=idx).value or "")[:50]) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)
        ws.freeze_panes = "A2"
    return buf.getvalue()

compare_excel = _compare_to_excel(df, compare_job)
st.download_button(
    "⬇️ Export Comparison to Excel",
    compare_excel,
    file_name=f"{compare_job}_comparison.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    key="export_comparison",
    use_container_width=True,
    type="primary",
)

# ── Footer actions ---------------------------------------------------------------
b1, b2 = st.columns(2)
with b1:
    if st.button("← Back to Job Rankings", use_container_width=True):
        safe_switch_page("pages/2_Job_Rankings.py")
with b2:
    if st.button("Clear comparison selection", type="secondary",
                 use_container_width=True):
        st.session_state.pop("compare_ids", None)
        st.session_state.pop("compare_job", None)
        st.rerun()
